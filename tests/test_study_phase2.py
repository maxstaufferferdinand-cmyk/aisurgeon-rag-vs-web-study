from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from aisurgeon_decentralised.evidence_contract import (
    EvidenceRecord,
    ExcludedEvidenceError,
    build_evidence_package,
)
from aisurgeon_decentralised.rag_core import RagCore, RagHit, RetrievalMode
from aisurgeon_decentralised.study_costs import calculate_cost
from aisurgeon_decentralised.study_exports import validate_study_workbooks
from aisurgeon_decentralised.study_model_verification import (
    EXPECTED_GPT55_SNAPSHOT,
    require_recent_model_verification,
)
from aisurgeon_decentralised.study_phase2 import (
    CORPUS_SNAPSHOT_ID,
    MAX_WEB_TOOL_CALLS,
    MODEL_CONFIGURATIONS,
    PRIMARY_RESULT_COUNT,
    STUDY_MAX_ESTIMATED_API_COST_USD,
    STUDY_OWNER_APPROVED_QUESTIONS_SHA256,
    SUPERSEDED_STUDY_COST_CEILING_USD,
    HumanQuestionFreezeRequired,
    StudyQuestion,
    build_randomization_manifest,
    load_frozen_questions,
    read_jsonl,
    sha256_file,
    validate_question_set,
)
from aisurgeon_decentralised.study_question_bank import (
    COVERED_DRAFTS,
    NOT_COVERED_DRAFTS,
)
from aisurgeon_decentralised.study_ratings import weighted_kappa
from aisurgeon_decentralised.study_responses import (
    ResponseUsage,
    StudyResponsesError,
    StudyStructuredAnswer,
    extract_web_provenance,
    response_json_schema,
)
from aisurgeon_decentralised.study_runner import (
    _attempt_from_error,
    pending_study_cells,
)
from aisurgeon_decentralised.study_statistics import (
    build_reproducibility_statistics,
    cluster_bootstrap_mean_ci,
)
from aisurgeon_decentralised.study_validators import (
    validate_rag_answer,
    validate_web_answer,
)

ROOT = Path(__file__).resolve().parents[1]
QUESTIONS_PATH = ROOT / "outputs/study_phase2/questions/question_candidates.jsonl"


def _questions() -> tuple[StudyQuestion, ...]:
    return tuple(
        StudyQuestion.model_validate(row) for row in read_jsonl(QUESTIONS_PATH)
    )


def _answer(source_ref: str, *, status: str = "supported") -> StudyStructuredAnswer:
    return StudyStructuredAnswer.model_validate(
        {
            "answer_status": status,
            "answer_text": "Eine belegte Testantwort.",
            "claims": [
                {
                    "claim_id": "c1",
                    "claim_text": "Eine belegte Testantwort.",
                    "claim_type": "fact",
                    "support_status": status,
                    "source_refs": [source_ref],
                }
            ],
            "recommendations": [],
            "limitations": [],
            "abstention_reason": None,
        }
    )


def _evidence(*, excluded: bool = False) -> EvidenceRecord:
    return EvidenceRecord(
        evidence_id="ru-test",
        corpus_snapshot_id=CORPUS_SNAPSHOT_ID,
        source_document_id="src-test",
        source_version_id="sv-test",
        document_name="Testleitlinie",
        version_label="1",
        source_status="final",
        source_role="guideline",
        source_authority="test",
        document_component="guideline",
        source_file_name="test.pdf",
        source_link="source_pdfs/test.pdf",
        exact_source_text="Eine belegte Testantwort.",
        pdf_pages_1based=(1,),
        eligibility_status="ineligible" if excluded else "eligible",
        retrieval_eligible=not excluded,
        answer_eligible=not excluded,
        excluded_by_policy=excluded,
        exclusion_reason="hcc_historical_change_table" if excluded else None,
    )


def test_question_bank_and_exports_have_exact_prespecified_counts() -> None:
    assert len(COVERED_DRAFTS) == 80
    assert len(NOT_COVERED_DRAFTS) == 20
    questions = _questions()
    validate_question_set(questions, require_human_freeze=False)
    assert len(questions) == 100
    assert sum(q.coverage_stratum == "covered_by_local_corpus" for q in questions) == 80
    assert (
        sum(q.coverage_stratum == "not_covered_by_local_corpus" for q in questions)
        == 20
    )


def test_no_phase1_development_question_is_reused() -> None:
    study_texts = {row.question_text for row in _questions()}
    development = {
        row["question_text"]
        for row in read_jsonl(
            ROOT / "outputs/retrieval_phase/vte_development/vte_questions.jsonl"
        )
    }
    assert not study_texts.intersection(development)


def test_design_has_two_models_two_arms_two_runs_and_800_unique_cells() -> None:
    questions = _questions()
    cells = build_randomization_manifest(questions, frozen=False)
    assert len(MODEL_CONFIGURATIONS) == 2
    assert {cell.system_arm for cell in cells} == {"WEB", "RAG"}
    assert {cell.repetition for cell in cells} == {
        "1_primary",
        "2_reproducibility",
    }
    assert len(cells) == PRIMARY_RESULT_COUNT == 800
    assert len({cell.run_id for cell in cells}) == 800
    counts: dict[tuple[str, str], int] = {}
    for cell in cells:
        key = (cell.model_config_id, cell.system_arm)
        counts[key] = counts.get(key, 0) + 1
    assert set(counts.values()) == {200}
    assert STUDY_MAX_ESTIMATED_API_COST_USD == 500.0
    assert SUPERSEDED_STUDY_COST_CEILING_USD == 400.0
    assert MAX_WEB_TOOL_CALLS == 6


def test_randomization_is_deterministic() -> None:
    questions = _questions()
    first = build_randomization_manifest(questions, frozen=False)
    second = build_randomization_manifest(questions, frozen=False)
    assert first == second


def test_human_freeze_gate_fails_closed(tmp_path: Path) -> None:
    with pytest.raises(HumanQuestionFreezeRequired):
        load_frozen_questions(tmp_path / "missing.jsonl")


def test_completed_and_failed_cells_are_not_reexecuted() -> None:
    cells = build_randomization_manifest(_questions(), frozen=False)
    recorded = [
        {"run_id": cells[0].run_id, "status": "complete"},
        {"run_id": cells[1].run_id, "status": "failed"},
    ]
    pending = pending_study_cells(cells, recorded)
    assert len(pending) == 798
    assert cells[0] not in pending
    assert cells[1] not in pending


def test_strict_schema_disallows_additional_properties_recursively() -> None:
    schema = response_json_schema()
    assert schema["strict"] is True

    def inspect(node: object) -> None:
        if isinstance(node, dict):
            if node.get("type") == "object":
                assert node.get("additionalProperties") is False
            for value in node.values():
                inspect(value)
        elif isinstance(node, list):
            for value in node:
                inspect(value)

    inspect(schema["schema"])


def test_rag_validator_accepts_allowlisted_backend_citation() -> None:
    row = _evidence()
    package = build_evidence_package(
        corpus_snapshot_id=CORPUS_SNAPSHOT_ID,
        evidence_ids=(row.evidence_id,),
        evidence_catalog={row.evidence_id: row},
    )
    hit = RagHit(
        evidence_id=row.evidence_id,
        rank=1,
        evidence_role="direct",
        source_role="guideline",
        source_status="final",
        source_document_id="src-test",
        pdf_pages_1based=(1,),
    )
    validated = validate_rag_answer(
        _answer("ru-test"), package=package, retrieval_hits=(hit,)
    )
    assert validated.validator_status == "accepted"
    assert validated.rendered_sources[0]["label"].startswith("Testleitlinie")


def test_rag_validator_rejects_invented_evidence_id() -> None:
    row = _evidence()
    package = build_evidence_package(
        corpus_snapshot_id=CORPUS_SNAPSHOT_ID,
        evidence_ids=(row.evidence_id,),
        evidence_catalog={row.evidence_id: row},
    )
    validated = validate_rag_answer(
        _answer("ru-invented"), package=package, retrieval_hits=()
    )
    assert validated.validator_status == "rejected"
    assert "unknown_or_not_allowlisted_evidence_id" in validated.issue_codes


def test_policy_excluded_history_cannot_enter_package() -> None:
    row = _evidence(excluded=True)
    with pytest.raises(ExcludedEvidenceError):
        build_evidence_package(
            corpus_snapshot_id=CORPUS_SNAPSHOT_ID,
            evidence_ids=(row.evidence_id,),
            evidence_catalog={row.evidence_id: row},
        )


def test_rag_validator_rejects_reverse_drug_bridge() -> None:
    row = _evidence()
    package = build_evidence_package(
        corpus_snapshot_id=CORPUS_SNAPSHOT_ID,
        evidence_ids=(row.evidence_id,),
        evidence_catalog={row.evidence_id: row},
    )
    hit = RagHit(
        evidence_id=row.evidence_id,
        rank=1,
        evidence_role="bridge_context",
        source_role="guideline",
        source_status="final",
        source_document_id="src-test",
        relation_types=("guideline_to_smpc",),
        seed_evidence_ids=("ru-smPC",),
    )
    validated = validate_rag_answer(
        _answer("ru-test"), package=package, retrieval_hits=(hit,)
    )
    assert validated.validator_status == "rejected"
    assert "invalid_or_reverse_drug_bridge_relation" in validated.issue_codes


def test_web_validator_is_separate_and_rejects_invented_url() -> None:
    consulted = ({"url": "https://official.example/guideline", "title": "Guideline"},)
    accepted = validate_web_answer(
        _answer("https://official.example/guideline"),
        consulted_sources=consulted,
        cited_sources=consulted,
    )
    assert accepted.validator_status == "accepted"
    rejected = validate_web_answer(
        _answer("https://invented.example/source"),
        consulted_sources=consulted,
        cited_sources=consulted,
    )
    assert rejected.validator_status == "rejected"
    assert "web_url_not_returned_by_current_search" in rejected.issue_codes


def test_web_validator_matches_only_known_tracking_variants() -> None:
    consulted = ({"url": "https://official.example/guideline?document=7"},)
    cited = (
        {
            "url": (
                "https://official.example/guideline?document=7&utm_source=openai"
            )
        },
    )
    accepted = validate_web_answer(
        _answer("https://official.example/guideline?document=7"),
        consulted_sources=consulted,
        cited_sources=cited,
    )
    assert accepted.validator_status == "accepted"


def test_rag_validator_rejects_unknown_recommendation_reference() -> None:
    row = _evidence()
    package = build_evidence_package(
        corpus_snapshot_id=CORPUS_SNAPSHOT_ID,
        evidence_ids=(row.evidence_id,),
        evidence_catalog={row.evidence_id: row},
    )
    answer = StudyStructuredAnswer.model_validate(
        {
            "answer_status": "supported",
            "answer_text": "Eine Empfehlung.",
            "claims": [
                {
                    "claim_id": "c1",
                    "claim_text": "Eine belegte Testantwort.",
                    "claim_type": "fact",
                    "support_status": "supported",
                    "source_refs": ["ru-test"],
                }
            ],
            "recommendations": [
                {
                    "recommendation_text": "Eine Empfehlung.",
                    "source_refs": ["ru-invented"],
                }
            ],
            "limitations": [],
            "abstention_reason": None,
        }
    )
    validated = validate_rag_answer(answer, package=package, retrieval_hits=())
    assert validated.validator_status == "downgraded"
    assert validated.recommendations[0].validator_status == "rejected"


def test_web_provenance_retains_access_and_annotation_metadata() -> None:
    raw = {
        "output": [
            {
                "type": "web_search_call",
                "action": {
                    "type": "search",
                    "query": "test",
                    "sources": [
                        {
                            "url": "https://official.example/guideline",
                            "title": "Guideline",
                            "snippet": "Public evidence excerpt",
                        }
                    ],
                },
            },
            {
                "type": "message",
                "content": [
                    {
                        "annotations": [
                            {
                                "type": "url_citation",
                                "url": "https://official.example/guideline",
                                "title": "Guideline",
                                "start_index": 2,
                                "end_index": 12,
                                "text": "Guideline",
                            }
                        ]
                    }
                ],
            },
        ]
    }
    consulted, cited, actions, calls = extract_web_provenance(raw)
    assert calls == 1
    assert actions[0]["query"] == "test"
    assert consulted[0]["accessed_at_utc"].endswith("Z")
    assert consulted[0]["content_hash"]
    assert cited[0]["citation_start_index"] == 2


def test_api_error_attempt_retains_telemetry_and_is_retryable() -> None:
    question = _questions()[0]
    cell = build_randomization_manifest(_questions(), frozen=False)[0]
    error = StudyResponsesError(
        "transient",
        status_code=None,
        error_code="network_or_sdk_error",
        client_request_id="client-test",
        api_wall_time_ms=12.5,
    )
    row = _attempt_from_error(
        experiment_id="test",
        cell=cell,
        question=question,
        error=error,
        attempt_number=1,
        started="2026-01-01T00:00:00Z",
        finished="2026-01-01T00:00:01Z",
        embedding_tokens=20,
        resources_before={"rss": 1},
        resources_after={"rss": 2},
    )
    assert row.retryable is True
    assert row.http_status is None
    assert row.query_embedding_tokens == 20
    assert row.total_estimated_cost_usd == pytest.approx(0.0000004)


def test_local_statistics_are_deterministic_without_llm_judge() -> None:
    assert weighted_kappa([0, 1, 2, 3], [0, 1, 2, 3]) == 1.0
    values = {"q1": -1.0, "q2": 0.5, "q3": 1.5}
    assert cluster_bootstrap_mean_ci(values, resamples=100, seed=7) == (
        cluster_bootstrap_mean_ci(values, resamples=100, seed=7)
    )
    pair_rows = [
        {
            "question_id": "q1",
            "model_config_id": "m",
            "system_arm": "RAG",
            "repetition": "1_primary",
            "coverage_stratum": "covered_by_local_corpus",
            "validated_system_answer": {
                "answer_status": "supported",
                "answer_text": "gleiche antwort",
                "claims": [],
                "recommendations": [],
            },
            "evidence_allowlist": ["ru-a"],
            "token_usage": {"total_tokens": 10},
            "cost": {"total_estimated_cost_usd": 0.1},
            "timing_ms": {"end_to_end": 10},
        },
        {
            "question_id": "q1",
            "model_config_id": "m",
            "system_arm": "RAG",
            "repetition": "2_reproducibility",
            "coverage_stratum": "covered_by_local_corpus",
            "validated_system_answer": {
                "answer_status": "supported",
                "answer_text": "gleiche antwort",
                "claims": [],
                "recommendations": [],
            },
            "evidence_allowlist": ["ru-a"],
            "token_usage": {"total_tokens": 11},
            "cost": {"total_estimated_cost_usd": 0.11},
            "timing_ms": {"end_to_end": 12},
        },
    ]
    rows, summary = build_reproducibility_statistics(pair_rows)
    assert rows[0]["answer_status_agreement"] is True
    assert rows[0]["answer_token_cosine_similarity"] == pytest.approx(1.0)
    assert summary["semantic_method"].startswith("local deterministic")


def test_rag_abstention_without_assertions_is_valid() -> None:
    package = build_evidence_package(
        corpus_snapshot_id=CORPUS_SNAPSHOT_ID,
        evidence_ids=(),
        evidence_catalog={},
    )
    answer = StudyStructuredAnswer.model_validate(
        {
            "answer_status": "no_validated_evidence",
            "answer_text": "Im lokalen Snapshot fehlt ausreichende Evidenz.",
            "claims": [],
            "recommendations": [],
            "limitations": ["Nur der lokale Snapshot wurde geprüft."],
            "abstention_reason": "Keine ausreichende lokale Evidenz.",
        }
    )
    validated = validate_rag_answer(answer, package=package, retrieval_hits=())
    assert validated.validator_status == "accepted"
    assert validated.answer_status == "no_validated_evidence"


def test_cost_calculation_does_not_double_count_reasoning_tokens() -> None:
    with_reasoning = ResponseUsage(
        input_tokens=1000,
        output_tokens=1000,
        reasoning_tokens=900,
        total_tokens=2000,
    )
    without_reasoning_detail = ResponseUsage(
        input_tokens=1000,
        output_tokens=1000,
        reasoning_tokens=0,
        total_tokens=2000,
    )
    first = calculate_cost(
        model="gpt-5.5-2026-04-23", usage=with_reasoning, web_search_calls=1
    )
    second = calculate_cost(
        model="gpt-5.5-2026-04-23",
        usage=without_reasoning_detail,
        web_search_calls=1,
    )
    assert first.total_estimated_cost_usd == second.total_estimated_cost_usd


def test_question_gold_ids_are_currently_policy_eligible() -> None:
    ids = {
        evidence_id
        for question in _questions()
        for evidence_id in question.expected_retrieval_unit_ids
    }
    from aisurgeon_decentralised.retrieval_database import connect

    with connect(ROOT) as connection, connection.cursor() as cursor:
        cursor.execute(
            "SELECT retrieval_unit_id FROM retrieval.eligible_retrieval_units "
            "WHERE corpus_snapshot_id=%s AND retrieval_unit_id=ANY(%s)",
            (CORPUS_SNAPSHOT_ID, list(ids)),
        )
        found = {row[0] for row in cursor.fetchall()}
    assert found == ids


def test_deterministic_fts_retrieval_repeats_identically() -> None:
    core = RagCore(root=ROOT, corpus_snapshot_id=CORPUS_SNAPSHOT_ID)
    text = _questions()[0].question_text
    first = core.retrieve(
        question=text, retrieval_mode=RetrievalMode.FTS, allow_embedding_api=False
    )
    second = core.retrieve(
        question=text, retrieval_mode=RetrievalMode.FTS, allow_embedding_api=False
    )
    assert first.evidence_ids == second.evidence_ids
    assert [hit.channel_ranks for hit in first.hits] == [
        hit.channel_ranks for hit in second.hits
    ]


def test_excel_exports_have_required_rows_and_blinding() -> None:
    assert validate_study_workbooks(root=ROOT)["status"] == "passed"
    rating = load_workbook(
        ROOT / "outputs/study_phase2/ratings/clinical_ratings_blinded.xlsx",
        read_only=True,
    )["RATINGS_BLINDED"]
    headers = {cell.value for cell in rating[1]}
    assert rating.max_row - 1 == 400
    assert "model" not in headers
    assert "system_arm" not in headers
    assert "sources" not in headers
    assert "run_id" not in headers


def test_provisional_questions_are_not_falsely_human_approved() -> None:
    questions = _questions()
    assert all(question.human_review_status == "pending" for question in questions)
    assert all(question.freeze_timestamp is None for question in questions)


def test_study_owner_freeze_is_exactly_100_and_substantively_unchanged() -> None:
    source = read_jsonl(QUESTIONS_PATH)
    frozen = read_jsonl(
        ROOT / "outputs/study_phase2/questions/study_questions_frozen.jsonl"
    )
    assert len(frozen) == 100
    metadata = {
        "confirmed_coverage_status",
        "confirmed_critical_errors",
        "confirmed_gold_sources_or_abstention",
        "confirmed_required_claims",
        "freeze_timestamp",
        "gold_standard_version",
        "human_review_status",
    }
    for original, approved in zip(source, frozen, strict=True):
        assert {k: v for k, v in original.items() if k not in metadata} == {
            k: v for k, v in approved.items() if k not in metadata
        }
        assert approved["human_review_status"] == (
            "study_owner_pre_freeze_approval"
        )


def test_active_cost_gate_is_500_and_runner_loaded_it() -> None:
    from aisurgeon_decentralised import study_runner

    manifest = json.loads(
        (
            ROOT / "outputs/study_phase2/manifest/study_manifest.json"
        ).read_text(encoding="utf-8")
    )
    price = json.loads(
        (ROOT / "outputs/study_phase2/manifest/price_table.json").read_text(
            encoding="utf-8"
        )
    )
    assert study_runner.STUDY_MAX_ESTIMATED_API_COST_USD == 500.0
    assert manifest["cost_ceiling_usd"] == 500.0
    assert manifest["superseded_cost_ceiling_usd"] == 400.0
    assert manifest["cost_counters_reset"] is False
    assert price["study_cost_ceiling_usd"] == 500.0
    assert price["supersedes_study_cost_ceiling_usd"] == 400.0


def test_owner_approval_has_no_invented_reviewers() -> None:
    approval = json.loads(
        (
            ROOT
            / "outputs/study_phase2/questions/"
            "study_owner_pre_freeze_approval.json"
        ).read_text(encoding="utf-8")
    )
    assert approval["approval_type"] == "study_owner_pre_freeze_approval"
    assert approval["question_candidates_sha256"] == (
        STUDY_OWNER_APPROVED_QUESTIONS_SHA256
    )
    assert approval["independent_question_set_review_completed"] is False
    assert approval["reviewer_name_recorded"] is False
    assert approval["signature_recorded"] is False


def test_official_model_verification_is_current_and_unambiguous() -> None:
    report = require_recent_model_verification(root=ROOT)
    assert report["official_sources_only"] is True
    assert report["detected_gpt55_dated_snapshots"] == [EXPECTED_GPT55_SNAPSHOT]
    assert report["detected_gpt56_sol_dated_snapshots"] == []
    assert report["frozen_gpt56_sol_request"] == "gpt-5.6-sol"


def test_study_owner_freeze_artifact_hashes_match_files() -> None:
    manifest = json.loads(
        (
            ROOT
            / "outputs/study_phase2/manifest/"
            "artifact_hashes_study_owner_freeze_v2_500usd.json"
        ).read_text(encoding="utf-8")
    )
    for artifact in manifest["artifacts"]:
        path = ROOT / artifact["path"]
        assert path.stat().st_size == artifact["bytes"]
        assert sha256_file(path) == artifact["sha256"]
