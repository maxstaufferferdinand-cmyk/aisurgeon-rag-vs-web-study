"""Versioned pre-freeze study-owner approval and cost-ceiling amendment."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .study_costs import PRICE_TABLE
from .study_phase2 import (
    PROTOCOL_VERSION,
    STUDY_MAX_ESTIMATED_API_COST_USD,
    STUDY_OWNER_APPROVED_QUESTIONS_SHA256,
    SUPERSEDED_STUDY_COST_CEILING_USD,
    read_jsonl,
    sha256_file,
    utc_now,
)
from .study_runner import build_pilot_cost_projection

ARCHIVE_RELATIVE = Path(
    "outputs/study_phase2/manifest/history/400usd_pre_owner_approval"
)
ARCHIVED_HASHES = {
    "STATISTICAL_ANALYSIS_PLAN_RAG_VS_WEB.md": (
        "e9c18efae911f1dde9d1cc8eadf8ff9e60920101dc53ae704146d04fbd506538"
    ),
    "STUDY_PROTOCOL_RAG_VS_WEB.md": (
        "47b6009cce76584054355542d26f5eea876ccf5f699264bfff6380d068342ad9"
    ),
    "artifact_hashes_pre_human_freeze.json": (
        "f8ca7934d925f1936e43264196b60bd875a566ccfb16c848c39ebdb1cac61a48"
    ),
    "development_cost_pilot_summary.json": (
        "f1aca5613954a12b880344eb919905fabf5438084d8db479aeb0f76c8bf295ed"
    ),
    "phase2_validation.json": (
        "d0a3f5e9a2b34ee046db61543c757cb665da119ef13fa153e8f9cebe84b8d3dd"
    ),
    "price_table.json": (
        "4363302530acdc69872e95c5b14974a172818168e67926459d0918a3fcd87613"
    ),
    "protocol_deviations.json": (
        "17138b1ce37a0e3484772b0c08f291d2f097943dc4d6f483987b06d6c1ed1b34"
    ),
    "study_manifest.json": (
        "41ed248b8542ccb35b15df6dbabc0c58e99241d079e9f7fcb9191f9b57c4cd3f"
    ),
    "technical_completion_report.json": (
        "43c2a7df5efa5a7096a49d3883e0d8689363eff5cd0a0f8d955635c4fdb8c098"
    ),
    "technical_completion_report.md": (
        "84ffe9fde76b97c3f82a6496c51c217ed6c58e10a14088b1713ecc46c6175862"
    ),
}


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(path)


def _verify_blank_review_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["01_FRAGEN_GOLD"]
    headers = [str(cell.value or "") for cell in sheet[1]]
    first_review = headers.index("Reviewer_A_Name")
    nonempty: list[str] = []
    question_ids: list[str] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        question_id = str(values[0] or "")
        if question_id:
            question_ids.append(question_id)
        for index, value in enumerate(values[first_review:], start=first_review):
            if value not in (None, ""):
                nonempty.append(f"{question_id}:{headers[index]}")
    if len(question_ids) != 100 or len(set(question_ids)) != 100:
        raise RuntimeError("question review workbook does not contain 100 unique rows")
    if nonempty:
        raise RuntimeError(
            "review workbook contains unexpected reviewer/adjudication content: "
            f"{nonempty[:10]}"
        )
    return {
        "path": str(path),
        "sha256": sha256_file(path),
        "question_rows": len(question_ids),
        "reviewer_and_adjudication_fields_intentionally_blank": True,
        "used_as_independent_review": False,
    }


def _update_compliance(base: Path) -> None:
    limitation = (
        "Pre-freeze amendment PD-001: the unchanged synthetic question/gold draft "
        "was approved by the study owner only; no independent clinical question-set "
        "review occurred. Independent blinded answer ratings remain mandatory."
    )
    for filename in ("REFINE_compliance.csv", "MI_CLEAR_LLM_compliance.csv"):
        path = base / "manifest" / filename
        with path.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
        for row in rows:
            requirement = str(row.get("requirement") or "").casefold()
            if any(
                marker in requirement
                for marker in (
                    "human review",
                    "reference standard",
                    "test-data independence",
                    "independent human question",
                    "blind primary clinical ratings",
                )
            ):
                row["pre_run_status"] = (
                    "implemented_with_study_owner_approval_limitation"
                )
            prior = str(row.get("notes") or "").strip()
            while limitation in prior:
                prior = prior.replace(limitation, "").strip()
            row["notes"] = f"{prior} {limitation}".strip()
        _write_csv(path, rows)


def apply_study_owner_prefreeze_amendment(*, root: Path) -> dict[str, Any]:
    root = root.resolve()
    base = root / "outputs/study_phase2"
    candidate = base / "questions/question_candidates.jsonl"
    provisional = base / "questions/question_gold_provisional.jsonl"
    for path in (candidate, provisional):
        if sha256_file(path) != STUDY_OWNER_APPROVED_QUESTIONS_SHA256:
            raise RuntimeError(
                f"study-owner-approved question input changed: {path}"
            )
    if candidate.read_bytes() != provisional.read_bytes():
        raise RuntimeError("candidate and provisional-gold bytes differ")
    if (base / "results/study_results.jsonl").is_file():
        raise RuntimeError("pre-freeze amendment cannot be applied after main results")

    archive = root / ARCHIVE_RELATIVE
    archive_evidence = []
    for name, expected in ARCHIVED_HASHES.items():
        path = archive / name
        if not path.is_file() or sha256_file(path) != expected:
            raise RuntimeError(f"400-USD historical archive mismatch: {path}")
        archive_evidence.append(
            {"path": str(path.relative_to(root)), "sha256": expected}
        )
    approval_path = base / "questions/study_owner_pre_freeze_approval.json"
    existing_approval = (
        json.loads(approval_path.read_text(encoding="utf-8"))
        if approval_path.is_file()
        else None
    )
    now = (
        str(existing_approval["recorded_at_utc"])
        if existing_approval
        else utc_now()
    )
    supersession = {
        "schema_version": "study-cost-ceiling-supersession-1.0.0",
        "recorded_at_utc": now,
        "historical_cost_ceiling_usd": SUPERSEDED_STUDY_COST_CEILING_USD,
        "historical_status": "superseded_pre_main_study",
        "superseded_by_cost_ceiling_usd": STUDY_MAX_ESTIMATED_API_COST_USD,
        "superseded_by_protocol_version": PROTOCOL_VERSION,
        "historical_artifacts": archive_evidence,
        "historical_files_modified": False,
    }
    _write_json(archive / "supersession_manifest.json", supersession)

    review_audit = _verify_blank_review_workbook(
        base / "questions/question_freeze_review.xlsx"
    )
    owner_approval = {
        "schema_version": "study-owner-prefreeze-approval-1.0.0",
        "approval_type": "study_owner_pre_freeze_approval",
        "recorded_at_utc": now,
        "approved_by_role": "study_owner",
        "reviewer_name_recorded": False,
        "signature_recorded": False,
        "comments_recorded": False,
        "approval_scope": (
            "exact unchanged 100-question candidate and provisional-gold files, "
            "including existing gold IDs and coverage assignments"
        ),
        "question_candidates_sha256": sha256_file(candidate),
        "question_gold_provisional_sha256": sha256_file(provisional),
        "question_count": len(read_jsonl(candidate)),
        "independent_question_set_review_completed": False,
        "independent_clinical_question_validation_claimed": False,
        "later_blinded_answer_accuracy_review_required": True,
        "review_workbook": review_audit,
        "status": "study_owner_pre_freeze_approval_recorded",
    }
    _write_json(approval_path, owner_approval)

    amendments = {
        "schema_version": "study-protocol-amendments-1.1.0",
        "deviations": [
            {
                "deviation_id": "PD-001",
                "type": "pre_freeze_protocol_amendment",
                "recorded_at_utc": now,
                "title": "Study-owner approval replaces independent question freeze",
                "previous_requirement": (
                    "two independent question reviewers and adjudication before freeze"
                ),
                "amended_requirement": "study_owner_pre_freeze_approval",
                "reason": "explicit study-owner directive before any main-study call",
                "question_or_gold_content_changed": False,
                "scientific_impact": (
                    "question/gold independence is absent; clinical accuracy cannot be "
                    "inferred until later independent blinded answer ratings"
                ),
                "mitigation": (
                    "retain all question hashes; report limitation; preserve mandatory "
                    "two-reviewer blinded answer assessment and citation audit"
                ),
            },
            {
                "deviation_id": "PD-002",
                "type": "pre_freeze_cost_ceiling_amendment",
                "recorded_at_utc": now,
                "title": "Cumulative Phase-2 cost ceiling increased to 500 USD",
                "previous_cost_ceiling_usd": SUPERSEDED_STUDY_COST_CEILING_USD,
                "amended_cost_ceiling_usd": STUDY_MAX_ESTIMATED_API_COST_USD,
                "reason": "explicit study-owner directive before main-study execution",
                "cost_counters_reset": False,
                "historical_artifacts_preserved": True,
            },
        ],
    }
    _write_json(base / "manifest/protocol_deviations.json", amendments)
    _write_json(base / "manifest/price_table.json", PRICE_TABLE)

    pilot_results = read_jsonl(
        base / "pilot/development_cost_pilot_results.jsonl"
    )
    pilot_summary = build_pilot_cost_projection(pilot_results, root=root)
    _write_json(
        base / "pilot/development_cost_pilot_summary.json", pilot_summary
    )
    _update_compliance(base)

    manifest_path = base / "manifest/study_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update(
        {
            "protocol_version": PROTOCOL_VERSION,
            "status": "STUDY_OWNER_PRE_FREEZE_APPROVAL_RECORDED",
            "main_study_calls_allowed": False,
            "cost_ceiling_usd": STUDY_MAX_ESTIMATED_API_COST_USD,
            "superseded_cost_ceiling_usd": SUPERSEDED_STUDY_COST_CEILING_USD,
            "cost_counters_reset": False,
            "cost_ceiling_supersession_manifest": str(
                (archive / "supersession_manifest.json").relative_to(root)
            ),
            "study_owner_pre_freeze_approval": owner_approval,
            "study_owner_pre_freeze_approval_sha256": sha256_file(approval_path),
            "question_set_independently_clinically_validated": False,
            "protocol_deviations": amendments["deviations"],
            "development_cost_pilot": {
                **(manifest.get("development_cost_pilot") or {}),
                "status": "complete",
                "summary": pilot_summary,
            },
        }
    )
    _write_json(manifest_path, manifest)
    report = {
        "schema_version": "phase2-owner-prefreeze-amendment-report-1.0.0",
        "created_at_utc": now,
        "status": "STUDY_OWNER_PRE_FREEZE_APPROVAL_RECORDED",
        "protocol_version": PROTOCOL_VERSION,
        "question_count": owner_approval["question_count"],
        "approved_question_sha256": STUDY_OWNER_APPROVED_QUESTIONS_SHA256,
        "question_or_gold_content_changed": False,
        "approval_basis": "study_owner_pre_freeze_approval",
        "independent_question_set_review_completed": False,
        "later_independent_blinded_answer_review_required": True,
        "cost_ceiling_usd": STUDY_MAX_ESTIMATED_API_COST_USD,
        "superseded_cost_ceiling_usd": SUPERSEDED_STUDY_COST_CEILING_USD,
        "cost_counters_reset": False,
        "conservative_total_projection_usd": pilot_summary[
            "conservative_total_projection_usd"
        ],
        "main_study_results_recorded": 0,
        "protocol_deviations": amendments["deviations"],
    }
    _write_json(base / "reports/technical_completion_report.json", report)
    report_md = """# Phase-2-Bericht nach Pre-Freeze-Amendment

Status: `STUDY_OWNER_PRE_FREEZE_APPROVAL_RECORDED`

Die exakt unveränderten 100 Frage-/Goldentwürfe wurden durch den Study Owner
freigegeben. Es fand keine unabhängige klinische Fragevalidierung statt;
spätere verblindete Antwortbewertungen bleiben verpflichtend.

Aktives kumulatives Kostenlimit: 500.00 USD. Die frühere 400-USD-Grenze ist als
historisch abgelöst archiviert; Vorbereitungs- und Pilotkosten wurden nicht
zurückgesetzt.

Protokolländerungen: `PD-001`, `PD-002`.
"""
    (base / "reports/technical_completion_report.md").write_text(
        report_md, encoding="utf-8", newline="\n"
    )
    return {
        "status": "STUDY_OWNER_PRE_FREEZE_APPROVAL_RECORDED",
        "protocol_version": PROTOCOL_VERSION,
        "approved_question_sha256": STUDY_OWNER_APPROVED_QUESTIONS_SHA256,
        "questions": owner_approval["question_count"],
        "independent_question_review": False,
        "cost_ceiling_usd": STUDY_MAX_ESTIMATED_API_COST_USD,
        "superseded_cost_ceiling_usd": SUPERSEDED_STUDY_COST_CEILING_USD,
        "pilot_and_preparation_cost_counters_preserved": True,
    }


__all__ = [
    "ARCHIVED_HASHES",
    "ARCHIVE_RELATIVE",
    "apply_study_owner_prefreeze_amendment",
]
