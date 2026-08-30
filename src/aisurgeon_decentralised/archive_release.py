"""Build and validate the redistribution-safe GitHub study archive.

The historical corpus and study artifacts remain untouched.  This module only
creates new, explicitly redacted derivatives under ``archive/`` and an
allowlist for the first Git commit.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from collections.abc import Iterable, Mapping, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SNAPSHOT_ID = "cs-f61b3d4e90089c1b890c23cb"
ARCHIVE_SCHEMA_VERSION = "github-archive-1.0.0"
ARCHIVE_ROOT_NAME = "archive"

SOURCE_METADATA: dict[str, dict[str, str]] = {
    "003-001l_S3_Prophylaxe-venoese-Thromboembolie-VTE_2026-04.pdf": {
        "title": "S3-Leitlinie Prophylaxe der venösen Thromboembolie (VTE)",
        "publisher": "AWMF / beteiligte Fachgesellschaften",
        "version": "4.1",
        "official_url": "https://register.awmf.org/de/leitlinien/detail/003-001",
    },
    "032-010OLl_Exokrines-Pankreaskarzinom_2025-06.pdf": {
        "title": "S3-Leitlinie Exokrines Pankreaskarzinom",
        "publisher": "Leitlinienprogramm Onkologie (AWMF, DKG, Deutsche Krebshilfe)",
        "version": "3.1",
        "official_url": "https://www.leitlinienprogramm-onkologie.de/leitlinien/pankreaskarzinom/",
    },
    "S3_LL_HCC_und_BCC_Konsultationsfassung_Langversion_6.01 (1).pdf": {
        "title": "Diagnostik und Therapie des Hepatozellulären Karzinoms und biliärer Karzinome",
        "publisher": "Leitlinienprogramm Onkologie (AWMF, DKG, Deutsche Krebshilfe)",
        "version": "Konsultationsfassung 6.01",
        "official_url": "https://www.leitlinienprogramm-onkologie.de/leitlinien/hcc-und-biliaere-karzinome/",
    },
    "5-fu-medac-50-mg-ml-injektionsloesung.pdf": {
        "title": "5-FU medac 50 mg/ml, Injektionslösung – Fachinformation",
        "publisher": "medac GmbH",
        "version": "Stand der Information 11.2024",
        "official_url": "https://www.medac.de/therapiegebiete/produkte/5-fu-medac",
    },
    "cisplatin-teva-r-1-mg-ml-konzentrat.pdf": {
        "title": "Cisplatin Teva 1 mg/ml Konzentrat – Fachinformation",
        "publisher": "Teva GmbH",
        "version": "Version 5; Stand Februar 2025",
        "official_url": "https://www.teva.de/produkte/details.html?tx_kfiexensioproductsdb_productsviewing%5Baction%5D=viewDetail&tx_kfiexensioproductsdb_productsviewing%5Bcontroller%5D=Products&tx_kfiexensioproductsdb_productsviewing%5Bpzn%5D=6559642",
    },
    "abraxane-epar-product-information_de.pdf": {
        "title": "Abraxane, INN-paclitaxel – EPAR product information",
        "publisher": "European Medicines Agency",
        "version": "SHA-256-gefrorene Fassung am Corpus Snapshot",
        "official_url": "https://www.ema.europa.eu/en/medicines/human/EPAR/abraxane",
    },
    "eliquis-epar-product-information_de.pdf": {
        "title": "Eliquis, INN-apixaban – EPAR product information",
        "publisher": "European Medicines Agency",
        "version": "SHA-256-gefrorene Fassung am Corpus Snapshot",
        "official_url": "https://www.ema.europa.eu/en/medicines/human/EPAR/eliquis",
    },
    "enhertu-epar-product-information_de.pdf": {
        "title": "Enhertu, INN-trastuzumab deruxtecan – EPAR product information",
        "publisher": "European Medicines Agency",
        "version": "SHA-256-gefrorene Fassung am Corpus Snapshot",
        "official_url": "https://www.ema.europa.eu/en/medicines/human/EPAR/enhertu",
    },
    "keytruda-epar-product-information_de.pdf": {
        "title": "Keytruda, INN-pembrolizumab – EPAR product information",
        "publisher": "European Medicines Agency",
        "version": "SHA-256-gefrorene Fassung am Corpus Snapshot",
        "official_url": "https://www.ema.europa.eu/en/medicines/human/EPAR/keytruda",
    },
    "lixiana-epar-product-information_de.pdf": {
        "title": "Lixiana, INN-edoxaban – EPAR product information",
        "publisher": "European Medicines Agency",
        "version": "SHA-256-gefrorene Fassung am Corpus Snapshot",
        "official_url": "https://www.ema.europa.eu/en/medicines/human/EPAR/lixiana",
    },
    "plavix-epar-product-information_de.pdf": {
        "title": "Plavix, INN-clopidogrel – EPAR product information",
        "publisher": "European Medicines Agency",
        "version": "SHA-256-gefrorene Fassung am Corpus Snapshot",
        "official_url": "https://www.ema.europa.eu/en/medicines/human/EPAR/plavix",
    },
    "xarelto-epar-product-information_de.pdf": {
        "title": "Xarelto, INN-rivaroxaban – EPAR product information",
        "publisher": "European Medicines Agency",
        "version": "SHA-256-gefrorene Fassung am Corpus Snapshot",
        "official_url": "https://www.ema.europa.eu/en/medicines/human/EPAR/xarelto",
    },
}

ROOT_FILES = {
    ".env.example",
    ".gitattributes",
    ".gitignore",
    ".python-version",
    "AGENTS.md",
    "CODEX_EXTRACTION_PROMPT.md",
    "DATA_AVAILABILITY.md",
    "PROVENANCE.md",
    "README.md",
    "REPOSITORY_CONTENTS.md",
    "REPRODUCIBILITY.md",
    "docker-compose.yml",
    "pyproject.toml",
    "uv.lock",
}

SAFE_OUTPUT_FILES = {
    # Corpus structure, hashes and aggregate QA; no extracted full text.
    f"outputs/knowledge_corpus/manifests/corpus_snapshots/{SNAPSHOT_ID}.json",
    "outputs/knowledge_corpus/schemas/canonical_record.schema.json",
    "outputs/knowledge_corpus/schemas/gemini_extraction_envelope.schema.json",
    "outputs/knowledge_corpus/schemas/retrieval_unit.schema.json",
    "outputs/knowledge_corpus/qa/citation_completeness.json",
    "outputs/knowledge_corpus/qa/coverage_report.json",
    "outputs/knowledge_corpus/qa/duplicate_report.json",
    "outputs/knowledge_corpus/qa/extraction_report.md",
    "outputs/knowledge_corpus/qa/final_validation.json",
    "outputs/knowledge_corpus/qa/numbering_gap_audit.csv",
    "outputs/knowledge_corpus/qa/numbering_gap_audit.json",
    "outputs/knowledge_corpus/qa/numbering_gap_audit.md",
    "outputs/knowledge_corpus/qa/preflight_results.json",
    "outputs/knowledge_corpus/qa/source_integrity_final.json",
    "outputs/knowledge_corpus/qa/targeted_repair_remaining.csv",
    "outputs/knowledge_corpus/qa/targeted_repair_report.json",
    "outputs/knowledge_corpus/qa/targeted_repair_report.md",
    "outputs/knowledge_corpus/retrieval/corpus_statistics.json",
    # Phase-1 bridge, validation, reports and public synthetic development data.
    "outputs/retrieval_phase/bridges/bridge_matrix.md",
    "outputs/retrieval_phase/bridges/bridge_qa.json",
    "outputs/retrieval_phase/bridges/smpc_guideline_bridge.csv",
    "outputs/retrieval_phase/bridges/smpc_guideline_bridge.jsonl",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/embeddings/text-embedding-3-small/embedding_validation.json",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/embeddings/text-embedding-3-small/full_report.json",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/embeddings/text-embedding-3-small/smoke_report.json",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/qa/database_rebuild.json",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/qa/phase_completion_report.json",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/qa/phase_completion_report.md",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/qa/retrieval_layer_validation.json",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/qa/retrieval_layer_validation.md",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/qa/semantic_retrieval_smoke.json",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/qa/structured_output_smoke.json",
    "outputs/retrieval_phase/evaluation/README.md",
    "outputs/retrieval_phase/evaluation/REVIEWER_GUIDE.md",
    "outputs/retrieval_phase/evaluation/adjudication.schema.json",
    "outputs/retrieval_phase/evaluation/annotation.schema.json",
    "outputs/retrieval_phase/evaluation/development_blind_questions.csv",
    "outputs/retrieval_phase/evaluation/development_blind_questions.jsonl",
    "outputs/retrieval_phase/evaluation/metrics_plan.json",
    "outputs/retrieval_phase/evaluation/sampling_manifest.json",
    "outputs/retrieval_phase/evaluation/test_blind_questions.csv",
    "outputs/retrieval_phase/evaluation/test_blind_questions.jsonl",
    "outputs/retrieval_phase/qa/cli_rag_phase1_completion.json",
    "outputs/retrieval_phase/qa/cli_rag_phase1_completion.md",
    "outputs/retrieval_phase/qa/cli_rag_phase1_report.md",
    "outputs/retrieval_phase/qa/rationale_relation_audit.json",
    "outputs/retrieval_phase/qa/rationale_relation_audit.md",
    "outputs/retrieval_phase/vte_development/benchmark_qa.json",
    "outputs/retrieval_phase/vte_development/determinism_qa.json",
    "outputs/retrieval_phase/vte_development/query_embedding_usage.json",
    "outputs/retrieval_phase/vte_development/response_cost_gate.json",
    "outputs/retrieval_phase/vte_development/response_summary.json",
    "outputs/retrieval_phase/vte_development/retrieval_evaluation.csv",
    "outputs/retrieval_phase/vte_development/retrieval_evaluation.jsonl",
    "outputs/retrieval_phase/vte_development/retrieval_metrics.json",
    "outputs/retrieval_phase/vte_development/vte_development_report.md",
    "outputs/retrieval_phase/vte_development/vte_questions.csv",
    "outputs/retrieval_phase/vte_development/vte_questions.jsonl",
    "outputs/retrieval_phase/vte_development/vte_questions.md",
    # Phase-2 immutable inputs, aggregate analyses and completion evidence.
    "outputs/study_phase2/analysis/development_cost_pilot_analysis.json",
    "outputs/study_phase2/analysis/question_summary.json",
    "outputs/study_phase2/analysis/rag_retrieval_metrics.csv",
    "outputs/study_phase2/analysis/rag_retrieval_metrics.json",
    "outputs/study_phase2/analysis/reproducibility.csv",
    "outputs/study_phase2/analysis/reproducibility.json",
    "outputs/study_phase2/analysis/resource_summary.csv",
    "outputs/study_phase2/analysis/resource_summary.json",
    "outputs/study_phase2/manifest/MI_CLEAR_LLM_compliance.csv",
    "outputs/study_phase2/manifest/REFINE_compliance.csv",
    "outputs/study_phase2/manifest/artifact_hashes_pre_human_freeze.json",
    "outputs/study_phase2/manifest/artifact_hashes_study_owner_freeze_v2_500usd.json",
    "outputs/study_phase2/manifest/artifact_hashes_technical_complete.json",
    "outputs/study_phase2/manifest/environment_manifest.json",
    "outputs/study_phase2/manifest/model_availability_verification.json",
    "outputs/study_phase2/manifest/price_table.json",
    "outputs/study_phase2/manifest/protocol_deviations.json",
    "outputs/study_phase2/manifest/randomization_manifest.csv",
    "outputs/study_phase2/manifest/retrieval_config.json",
    "outputs/study_phase2/manifest/web_search_config.json",
    "outputs/study_phase2/manifest/history/400usd_pre_owner_approval/STATISTICAL_ANALYSIS_PLAN_RAG_VS_WEB.md",
    "outputs/study_phase2/manifest/history/400usd_pre_owner_approval/STUDY_PROTOCOL_RAG_VS_WEB.md",
    "outputs/study_phase2/manifest/history/400usd_pre_owner_approval/price_table.json",
    "outputs/study_phase2/manifest/history/400usd_pre_owner_approval/protocol_deviations.json",
    "outputs/study_phase2/manifest/history/400usd_pre_owner_approval/supersession_manifest.json",
    "outputs/study_phase2/pilot/deterministic_revalidation.json",
    "outputs/study_phase2/pilot/development_cost_pilot_resource_summary.csv",
    "outputs/study_phase2/pilot/development_cost_pilot_summary.json",
    "outputs/study_phase2/prompts/COMMON_TASK_v1.txt",
    "outputs/study_phase2/prompts/RESPONSE_SCHEMA_v1.json",
    "outputs/study_phase2/prompts/SOURCE_POLICY_RAG_v1.txt",
    "outputs/study_phase2/prompts/SOURCE_POLICY_WEB_v1.txt",
    "outputs/study_phase2/questions/question_freeze_review.xlsx",
    "outputs/study_phase2/questions/study_questions_frozen.jsonl",
    "outputs/study_phase2/qa/ruff_phase2.json",
    "outputs/study_phase2/reports/technical_completion_report.md",
}

FORBIDDEN_TRACKED_PREFIXES = (
    "source_pdfs/",
    "data/canonical/",
    "data/extraction_runs/",
    "outputs/knowledge_corpus/canonical/",
    "outputs/knowledge_corpus/checkpoints/",
    "outputs/knowledge_corpus/manifests/source_manifest.json",
    "outputs/knowledge_corpus/retrieval/embedding_input.jsonl",
    "outputs/knowledge_corpus/retrieval/retrieval_units",
    "outputs/retrieval_phase/cs-35",
    "outputs/retrieval_phase/cs-d2",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/provenance/",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/query_embeddings/",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/embeddings/text-embedding-3-small/full/",
    f"outputs/retrieval_phase/{SNAPSHOT_ID}/embeddings/text-embedding-3-small/smoke/",
    "outputs/study_phase2/results/",
    "outputs/study_phase2/excel/",
    "outputs/study_phase2/ratings/",
)

PRUNED_LOCAL_DIRECTORIES = {
    ".git": ("git_internal", "Git object database and local configuration are never repository payload."),
    ".venv": ("generated_environment_or_cache", "Virtual environment is reproduced from uv.lock."),
    ".pytest_cache": ("generated_environment_or_cache", "Pytest cache is locally generated."),
    ".ruff_cache": ("generated_environment_or_cache", "Ruff cache is locally generated."),
    ".retrieval-locks": ("generated_environment_or_cache", "Local process locks are not reproducibility artifacts."),
    "__pycache__": ("generated_environment_or_cache", "Python bytecode cache is locally generated."),
}

SECRET_PATTERNS = {
    "openai_key": re.compile(r"\bsk-(?:proj|svcacct|admin)-[A-Za-z0-9_-]{20,}\b"),
    "legacy_openai_key": re.compile(r"\bsk-[A-Za-z0-9]{48}\b"),
    "github_token": re.compile(r"\b(?:gh[pousr]_[A-Za-z0-9_]{20,}|github_pat_[A-Za-z0-9_]{20,})\b"),
    "private_key": re.compile(r"-----BEGIN (?:RSA |EC |OPENSSH |DSA )?PRIVATE KEY-----"),
    "authorization_value": re.compile(r"Authorization\s*[:=]\s*['\"]?Bearer\s+[A-Za-z0-9._~+/-]{16,}", re.IGNORECASE),
    "credentialed_dsn": re.compile(r"postgres(?:ql)?://[^\s:/]+:[^\s@]+@", re.IGNORECASE),
}
ABSOLUTE_USER_PATH = re.compile(
    r"(?:/home/[A-Za-z0-9._-]+/|/Users/[A-Za-z0-9._-]+/|[A-Za-z]:\\Users\\[^\\]+\\)"
)
KNOWN_LOCAL_SECRET_FILES = {".env.retrieval", ".env"}


def utc_now() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(dict(row), ensure_ascii=False, sort_keys=True) + "\n")


def _csv_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if value is None:
        return ""
    return value


def write_csv(path: Path, rows: Sequence[Mapping[str, Any]], fields: Sequence[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = sorted({key for row in rows for key in row})
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fields), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _csv_value(row.get(key)) for key in fields})


def sanitize_string(value: str, root: Path) -> str:
    value = value.replace(str(root.resolve()), "${PROJECT_ROOT}")
    value = re.sub(r"/home/[A-Za-z0-9._-]+", "${USER_HOME}", value)
    value = re.sub(r"/Users/[A-Za-z0-9._-]+", "${USER_HOME}", value)
    value = re.sub(r"[A-Za-z]:\\Users\\[^\\]+", "${USER_HOME}", value)
    for pattern in SECRET_PATTERNS.values():
        value = pattern.sub("[REDACTED_SECRET]", value)
    return value


def sanitize_value(value: Any, root: Path) -> Any:
    if isinstance(value, str):
        return sanitize_string(value, root)
    if isinstance(value, list):
        return [sanitize_value(item, root) for item in value]
    if isinstance(value, dict):
        return {key: sanitize_value(item, root) for key, item in value.items()}
    return value


def pseudonymize(identifier: str | None, prefix: str) -> str | None:
    if not identifier:
        return None
    return f"{prefix}-{hashlib.sha256(identifier.encode('utf-8')).hexdigest()[:24]}"


def _safe_web_source(source: Mapping[str, Any], root: Path) -> dict[str, Any]:
    allowed = (
        "type",
        "url",
        "title",
        "publisher",
        "published_at",
        "accessed_at_utc",
        "content_hash",
        "citation_start_index",
        "citation_end_index",
    )
    return {key: sanitize_value(source.get(key), root) for key in allowed if key in source}


def sanitize_study_result(row: Mapping[str, Any], root: Path) -> dict[str, Any]:
    safe = sanitize_value(dict(row), root)
    safe["attempt_ids"] = [
        pseudonymize(str(item), "attempt") for item in row.get("attempt_ids", [])
    ]
    resources = row.get("local_resources") or {}
    safe["local_resources"] = {
        key: sanitize_value(resources[key], root)
        for key in ("before_retrieval", "after_cell")
        if key in resources
    }
    safe["web_sources_consulted"] = [
        _safe_web_source(item, root) for item in row.get("web_sources_consulted", [])
    ]
    safe["web_sources_cited"] = [
        _safe_web_source(item, root) for item in row.get("web_sources_cited", [])
    ]
    actions: list[dict[str, Any]] = []
    for action in row.get("web_search_actions", []):
        actions.append(
            {
                key: sanitize_value(action.get(key), root)
                for key in ("type", "query", "queries")
                if key in action
            }
            | {"source_count": len(action.get("sources") or [])}
        )
    safe["web_search_actions"] = actions
    return safe


def sanitize_attempt(row: Mapping[str, Any], root: Path) -> dict[str, Any]:
    drop = {
        "client_request_id",
        "response_id",
        "x_request_id",
        "rate_limit_headers",
        "local_resources_before",
        "local_resources_after",
    }
    safe = {
        key: sanitize_value(value, root)
        for key, value in row.items()
        if key not in drop and key != "attempt_id"
    }
    safe["attempt_id"] = pseudonymize(str(row.get("attempt_id") or ""), "attempt")
    safe["operational_identifiers_redacted"] = True
    return safe


def build_source_manifest(root: Path, archive_root: Path) -> dict[str, Any]:
    snapshot_path = root / f"outputs/knowledge_corpus/manifests/corpus_snapshots/{SNAPSHOT_ID}.json"
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    if snapshot.get("retrieval_unit_count") != 4469 or snapshot.get("source_count") != 12:
        raise RuntimeError("unexpected corpus snapshot counts")
    records: list[dict[str, Any]] = []
    for source in snapshot["source_pdfs"]:
        filename = Path(source["relative_path"]).name
        metadata = SOURCE_METADATA.get(filename)
        if metadata is None:
            raise RuntimeError(f"missing archive source metadata for {filename}")
        local_path = root / source["relative_path"]
        if not local_path.is_file() or sha256_file(local_path) != source["source_sha256"]:
            raise RuntimeError(f"source integrity mismatch for {filename}")
        status = "consultation_draft" if "HCC_und_BCC" in filename else (
            "final" if source["source_id"].startswith(("src-003", "src-032")) else "current_at_snapshot"
        )
        records.append(
            {
                "source_document_id": source["source_id"],
                "document_title": metadata["title"],
                "publisher": metadata["publisher"],
                "version": metadata["version"],
                "document_kind": "guideline" if source["source_id"].startswith(("src-003", "src-032", "src-s3")) else "medicinal_product_information",
                "source_status": status,
                "official_retrieval_url": metadata["official_url"],
                "original_retrieval_date": "2026-08-13",
                "source_sha256": source["source_sha256"],
                "page_count": source["page_count"],
                "file_size_bytes": source["file_size_bytes"],
                "original_file_name": filename,
                "corpus_snapshot_id": SNAPSHOT_ID,
                "redistribution": "excluded_source_document",
                "rebuild_instruction": "Download from the official page, verify SHA-256, place under source_pdfs/, then follow REPRODUCIBILITY.md.",
            }
        )
    records.sort(key=lambda item: item["original_file_name"])
    payload = {
        "schema_version": "archive-source-manifest-1.0.0",
        "corpus_snapshot_id": SNAPSHOT_ID,
        "source_count": len(records),
        "page_count": sum(item["page_count"] for item in records),
        "sources": records,
        "rights_note": "Source PDFs and extracted full text are not redistributed. Consult each publisher's current terms before local download or reuse.",
    }
    write_json(archive_root / "corpus/source_manifest.json", payload)
    write_csv(archive_root / "corpus/source_manifest.csv", records)
    write_json(
        archive_root / "corpus/corpus_snapshot_manifest.json",
        sanitize_value(snapshot, root),
    )
    return payload


def _result_csv_rows(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        raw = row.get("raw_model_answer") or {}
        validated = row.get("validated_system_answer") or {}
        token = row.get("token_usage") or {}
        timing = row.get("timing_ms") or {}
        cost = row.get("cost") or {}
        retrieval = row.get("retrieval") or {}
        sources = row.get("backend_rendered_sources") or row.get("web_sources_cited") or []
        result.append(
            {
                "run_id": row.get("run_id"),
                "question_id": row.get("question_id"),
                "coverage_stratum": row.get("coverage_stratum"),
                "model_config_id": row.get("model_config_id"),
                "requested_model": row.get("requested_model"),
                "returned_model": row.get("returned_model"),
                "reasoning_effort": row.get("reasoning_effort"),
                "system_arm": row.get("system_arm"),
                "repetition": row.get("repetition"),
                "question_text": row.get("question_text"),
                "answer_status": validated.get("answer_status") or raw.get("answer_status"),
                "validated_answer": validated.get("answer_text"),
                "raw_answer": raw.get("answer_text"),
                "sources": sources,
                "input_tokens": token.get("input_tokens"),
                "cached_input_tokens": token.get("cached_input_tokens"),
                "output_tokens": token.get("output_tokens"),
                "reasoning_tokens": token.get("reasoning_tokens"),
                "total_tokens": token.get("total_tokens"),
                "query_embedding_tokens": token.get("query_embedding_tokens"),
                "web_search_calls": len(row.get("web_search_actions") or []),
                "retrieval_time_ms": retrieval.get("retrieval_time_ms"),
                "api_wall_time_ms": timing.get("api_wall"),
                "end_to_end_time_ms": timing.get("end_to_end"),
                "estimated_cost_usd": cost.get("total_estimated_cost_usd"),
                "reconciled_cost_usd": cost.get("reconciled_cost_usd"),
                "validator_status": row.get("validator_status"),
                "error_or_retry": row.get("error_code") or row.get("retry_count"),
                "corpus_snapshot_id": row.get("corpus_snapshot_id"),
                "question_hash": row.get("question_hash"),
                "prompt_hashes": row.get("prompt_hashes"),
                "response_schema_hash": row.get("response_schema_hash"),
                "retrieval_config_hash": row.get("retrieval_config_hash"),
                "web_config_hash": row.get("web_config_hash"),
            }
        )
    return result


def _claims_source_rows(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        validated = row.get("validated_system_answer") or {}
        for claim in validated.get("claims") or []:
            refs = claim.get("source_refs") or []
            if not refs:
                refs = [None]
            for ref in refs:
                output.append(
                    {
                        "run_id": row.get("run_id"),
                        "question_id": row.get("question_id"),
                        "model_config_id": row.get("model_config_id"),
                        "system_arm": row.get("system_arm"),
                        "repetition": row.get("repetition"),
                        "answer_status": validated.get("answer_status"),
                        "claim_id": claim.get("claim_id"),
                        "claim_type": claim.get("claim_type"),
                        "claim_text": claim.get("claim_text"),
                        "support_status": claim.get("support_status"),
                        "source_ref": ref,
                        "validator_status": row.get("validator_status"),
                    }
                )
    return output


def build_redacted_machine_results(root: Path, archive_root: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    raw_results = read_jsonl(root / "outputs/study_phase2/results/study_results.jsonl")
    raw_attempts = read_jsonl(root / "outputs/study_phase2/results/api_attempts.jsonl")
    if len(raw_results) != 800 or len({row["run_id"] for row in raw_results}) != 800:
        raise RuntimeError("expected exactly 800 unique study results")
    if len(raw_attempts) < 800:
        raise RuntimeError("attempt ledger is incomplete")
    safe_results = [sanitize_study_result(row, root) for row in raw_results]
    safe_attempts = [sanitize_attempt(row, root) for row in raw_attempts]
    result_dir = archive_root / "study_phase2/results"
    write_jsonl(result_dir / "study_results_redacted.jsonl", safe_results)
    write_csv(result_dir / "study_results_redacted.csv", _result_csv_rows(safe_results))
    write_jsonl(result_dir / "api_attempts_redacted.jsonl", safe_attempts)
    write_csv(result_dir / "api_attempts_redacted.csv", safe_attempts)
    claims = _claims_source_rows(safe_results)
    write_jsonl(result_dir / "claims_sources_redacted.jsonl", claims)
    write_csv(result_dir / "claims_sources_redacted.csv", claims)

    vte_source = root / "outputs/retrieval_phase/vte_development/response_runs_validated.jsonl"
    if vte_source.is_file():
        vte_rows = [sanitize_value(row, root) for row in read_jsonl(vte_source)]
        for row in vte_rows:
            for key in ("x_request_id", "response_id", "client_request_id", "rate_limit_headers"):
                row.pop(key, None)
        write_jsonl(archive_root / "phase1/vte_response_runs_redacted.jsonl", vte_rows)
    return safe_results, safe_attempts


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WEB_FILL = PatternFill("solid", fgColor="FCE4D6")
RAG_FILL = PatternFill("solid", fgColor="DDEBF7")
GPT55_FILL = PatternFill("solid", fgColor="E2F0D9")
GPT56_FILL = PatternFill("solid", fgColor="E4DFEC")


def _sheet_from_rows(
    workbook: Workbook,
    name: str,
    rows: Sequence[Mapping[str, Any]],
    *,
    fields: Sequence[str] | None = None,
    table_name: str | None = None,
) -> None:
    ws = workbook.create_sheet(name)
    if fields is None:
        fields = list(rows[0]) if rows else ["Hinweis"]
    ws.append(list(fields))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([_csv_value(row.get(field)) for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_index, field in enumerate(fields, 1):
        values = [str(ws.cell(row=i, column=col_index).value or "") for i in range(1, min(ws.max_row, 80) + 1)]
        width = min(max(12, len(str(field)) + 2, max((len(value) for value in values), default=0) + 1), 70)
        if any(term in str(field).lower() for term in ("frage", "antwort", "claim", "quelle", "json", "text")):
            width = max(width, 36)
        ws.column_dimensions[get_column_letter(col_index)].width = width
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if rows and table_name:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _flatten_scalars(value: Any, prefix: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if isinstance(value, dict):
        for key in sorted(value):
            child = f"{prefix}.{key}" if prefix else key
            rows.extend(_flatten_scalars(value[key], child))
    elif isinstance(value, list):
        rows.append({"Kennzahl": prefix, "Wert": json.dumps(value, ensure_ascii=False, sort_keys=True)})
    else:
        rows.append({"Kennzahl": prefix, "Wert": value})
    return rows


def _rating_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(cell.value or f"Spalte_{index}") for index, cell in enumerate(ws[1], 1)]
    rows = [dict(zip(headers, values, strict=True)) for values in ws.iter_rows(min_row=2, values_only=True) if any(value is not None for value in values)]
    wb.close()
    return rows


def _scrub_workbook(source: Path, target: Path, root: Path) -> None:
    workbook = load_workbook(source)
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = sanitize_string(cell.value, root)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    workbook.close()


def _excel_result_rows(results: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in _result_csv_rows(results):
        rows.append(
            {
                "Frage_ID": row["question_id"],
                "Corpus_Abdeckung": row["coverage_stratum"],
                "Modell": row["requested_model"],
                "Reasoning": row["reasoning_effort"],
                "System_WEB_oder_RAG": row["system_arm"],
                "Run": row["repetition"],
                "Frage": row["question_text"],
                "Antwortstatus": row["answer_status"],
                "Validierte_Antwort": row["validated_answer"],
                "Rohantwort": row["raw_answer"],
                "Quellen": row["sources"],
                "Input_Tokens": row["input_tokens"],
                "Cached_Input_Tokens": row["cached_input_tokens"],
                "Output_Tokens": row["output_tokens"],
                "Reasoning_Tokens": row["reasoning_tokens"],
                "Total_Tokens": row["total_tokens"],
                "Web_Search_Aufrufe": row["web_search_calls"],
                "Retrievalzeit_ms": row["retrieval_time_ms"],
                "API_Zeit_ms": row["api_wall_time_ms"],
                "End_to_End_Zeit_ms": row["end_to_end_time_ms"],
                "Geschaetzte_Kosten_USD": row["estimated_cost_usd"],
                "Abgeglichene_Kosten_USD": row["reconciled_cost_usd"],
                "Validatorstatus": row["validator_status"],
                "Fehler_oder_Retry": row["error_or_retry"],
                "Run_ID": row["run_id"],
                "Question_Hash": row["question_hash"],
                "Corpus_Snapshot": row["corpus_snapshot_id"],
                "Returned_Model": row["returned_model"],
                "Prompt_Hashes": row["prompt_hashes"],
                "Schema_Hash": row["response_schema_hash"],
                "Retrieval_Config_Hash": row["retrieval_config_hash"],
                "Web_Config_Hash": row["web_config_hash"],
            }
        )
    return rows


def build_excel_exports(root: Path, archive_root: Path, results: Sequence[Mapping[str, Any]], attempts: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    excel_dir = archive_root / "study_phase2/excel"
    excel_dir.mkdir(parents=True, exist_ok=True)
    result_rows = _excel_result_rows(results)
    question_rows = read_jsonl(root / "outputs/study_phase2/questions/study_questions_frozen.jsonl")
    question_flat = [
        {key: _csv_value(value) for key, value in row.items()} for row in question_rows
    ]
    manifest_rows = [
        {"Merkmal": "Archivschema", "Wert": ARCHIVE_SCHEMA_VERSION},
        {"Merkmal": "Corpus Snapshot", "Wert": SNAPSHOT_ID},
        {"Merkmal": "Retrievaleinheiten", "Wert": 4469},
        {"Merkmal": "Corpus-Embeddings (nicht verteilt)", "Wert": 4469},
        {"Merkmal": "Fragen", "Wert": 100},
        {"Merkmal": "Studienzellen", "Wert": 800},
        {"Merkmal": "Modelle", "Wert": "gpt-5.5-2026-04-23/medium; gpt-5.6-sol/high"},
        {"Merkmal": "Status", "Wert": "TECHNICAL_STUDY_COMPLETE_CLINICAL_RATING_PENDING"},
        {"Merkmal": "Redaktion", "Wert": "Request-/Response-IDs, Rate-Limit-Header, lokale Pfade, Volltext-Evidence und Vektoren entfernt"},
    ]
    readme_rows = [
        {"Thema": "Zweck", "Erklärung": "Reproduzierbarer, redigierter Export der technischen AISurgeon-RAG-vs-WEB-Studie."},
        {"Thema": "WEB/RAG", "Erklärung": "WEB nutzt verpflichtende Live-Websuche; RAG nutzt ausschließlich das lokale Evidence Package."},
        {"Thema": "Run 1/2", "Erklärung": "Run 1 ist der primäre Lauf, Run 2 der Reproduzierbarkeitslauf; es wurde keine bessere Antwort nachträglich ausgewählt."},
        {"Thema": "Kosten", "Erklärung": "Geschätzte Kosten stammen aus der eingefrorenen Preistabelle; abgeglichene Kosten bleiben leer, wenn keine Admin-Abrechnung vorlag."},
        {"Thema": "Validierung", "Erklärung": "Technische Provenienzvalidierung ist keine unabhängige klinische Richtigkeitsvalidierung."},
        {"Thema": "Farben", "Erklärung": "WEB orange, RAG blau; GPT-5.5 grün, GPT-5.6 Sol violett. Farben dienen nur der Lesbarkeit."},
        {"Thema": "Filter", "Erklärung": "Kopfzeile auswählen und Excel-AutoFilter nach Frage, Modell, Arm, Run, Status oder Kosten verwenden."},
        {"Thema": "Datenschutz", "Erklärung": "Keine Patientendaten; operative Provider-IDs und lokale Benutzerpfade wurden für GitHub entfernt."},
    ]
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet_from_rows(workbook, "00_README", readme_rows, table_name="ArchiveReadme")
    _sheet_from_rows(workbook, "01_MANIFEST", manifest_rows, table_name="ArchiveManifest")
    _sheet_from_rows(workbook, "02_QUESTIONS_GOLD", question_flat, table_name="QuestionsGold")
    _sheet_from_rows(workbook, "03_ALL_RESULTS", result_rows, table_name="AllResults")
    _sheet_from_rows(workbook, "04_API_ATTEMPTS", attempts, table_name="ApiAttempts")
    _sheet_from_rows(workbook, "05_CLAIMS_SOURCES", _claims_source_rows(results), table_name="ClaimsSources")
    _sheet_from_rows(workbook, "06_RESOURCE_SUMMARY", _read_csv(root / "outputs/study_phase2/analysis/resource_summary.csv"), table_name="ResourceSummary")
    _sheet_from_rows(workbook, "07_REPRODUCIBILITY", _read_csv(root / "outputs/study_phase2/analysis/reproducibility.csv"), table_name="Reproducibility")
    ratings = _rating_rows(root / "outputs/study_phase2/ratings/clinical_ratings_blinded.xlsx")
    _sheet_from_rows(workbook, "08_RATINGS_BLINDED", ratings, table_name="RatingsBlinded")
    blinded_key = next((key for key in ratings[0] if "blinded" in key.lower() and "id" in key.lower()), None) if ratings else None
    adjudication = [
        {
            "blinded_response_id": row.get(blinded_key) if blinded_key else index,
            "reviewer_a": "",
            "reviewer_b": "",
            "adjudicated_rating": "",
            "adjudication_comment": "",
        }
        for index, row in enumerate(ratings, 1)
    ]
    _sheet_from_rows(workbook, "09_ADJUDICATION", adjudication, table_name="Adjudication")
    report = json.loads((root / "outputs/study_phase2/reports/technical_completion_report.json").read_text(encoding="utf-8"))
    _sheet_from_rows(workbook, "10_FINAL_STATISTICS", _flatten_scalars(sanitize_value(report, root)), table_name="FinalStatistics")
    compliance: list[dict[str, Any]] = []
    for framework, filename in (
        ("REFINE", "REFINE_compliance.csv"),
        ("MI-CLEAR-LLM", "MI_CLEAR_LLM_compliance.csv"),
    ):
        for row in _read_csv(root / f"outputs/study_phase2/manifest/{filename}"):
            compliance.append({"Framework": framework} | row)
    _sheet_from_rows(workbook, "11_REFINE_MI_CLEAR", compliance, table_name="Compliance")
    master_path = excel_dir / "AISurgeon_RAG_vs_WEB_study_master.xlsx"
    workbook.save(master_path)
    workbook.close()

    arms = {
        "GPT55_MEDIUM_WEB.xlsx": ("gpt55_medium", "WEB"),
        "GPT55_MEDIUM_RAG.xlsx": ("gpt55_medium", "RAG"),
        "GPT56_SOL_HIGH_WEB.xlsx": ("gpt56_sol_high", "WEB"),
        "GPT56_SOL_HIGH_RAG.xlsx": ("gpt56_sol_high", "RAG"),
    }
    for filename, (model_config, arm) in arms.items():
        rows = [
            item for item, raw in zip(result_rows, results, strict=True)
            if raw.get("model_config_id") == model_config and raw.get("system_arm") == arm
        ]
        if len(rows) != 200:
            raise RuntimeError(f"{filename}: expected 200 rows, got {len(rows)}")
        wb = Workbook()
        ws = wb.active
        ws.title = "00_README"
        ws.append(["Hinweis", f"Redigierter Archivexport: {model_config}, Arm {arm}, genau 200 Studienzellen."])
        ws.freeze_panes = "A2"
        _sheet_from_rows(wb, "RESULTS", rows, table_name=re.sub(r"[^A-Za-z0-9]", "", filename)[:25])
        wb.save(excel_dir / filename)
        wb.close()

    rating_dir = archive_root / "study_phase2/ratings"
    _scrub_workbook(
        root / "outputs/study_phase2/ratings/clinical_ratings_blinded.xlsx",
        rating_dir / "clinical_ratings_blinded.xlsx",
        root,
    )
    _scrub_workbook(
        root / "outputs/study_phase2/ratings/citation_audit.xlsx",
        rating_dir / "citation_audit.xlsx",
        root,
    )
    return validate_excel_exports(archive_root)


def validate_excel_exports(archive_root: Path) -> dict[str, Any]:
    excel_dir = archive_root / "study_phase2/excel"
    master = load_workbook(excel_dir / "AISurgeon_RAG_vs_WEB_study_master.xlsx", read_only=False, data_only=False)
    expected_sheets = [
        "00_README", "01_MANIFEST", "02_QUESTIONS_GOLD", "03_ALL_RESULTS",
        "04_API_ATTEMPTS", "05_CLAIMS_SOURCES", "06_RESOURCE_SUMMARY",
        "07_REPRODUCIBILITY", "08_RATINGS_BLINDED", "09_ADJUDICATION",
        "10_FINAL_STATISTICS", "11_REFINE_MI_CLEAR",
    ]
    issues: list[str] = []
    if master.sheetnames != expected_sheets:
        issues.append("master_sheet_names")
    result_ws = master["03_ALL_RESULTS"]
    headers = [cell.value for cell in result_ws[1]]
    run_col = headers.index("Run_ID") + 1
    run_ids = [result_ws.cell(row=row, column=run_col).value for row in range(2, result_ws.max_row + 1)]
    if len(run_ids) != 800 or len(set(run_ids)) != 800:
        issues.append("master_result_rows_or_run_ids")
    if result_ws.freeze_panes != "A2" or not result_ws.auto_filter.ref:
        issues.append("master_filter_or_freeze")
    formula_count = sum(
        1 for ws in master.worksheets for row in ws.iter_rows() for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    if formula_count:
        issues.append("formula_cells_present")
    master.close()
    files: list[dict[str, Any]] = []
    for path in sorted(excel_dir.glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=False)
        row_count = 800 if path.name.startswith("AISurgeon") else wb["RESULTS"].max_row - 1
        if not path.name.startswith("AISurgeon") and row_count != 200:
            issues.append(f"{path.name}:row_count")
        files.append({"file": path.relative_to(archive_root).as_posix(), "bytes": path.stat().st_size, "data_rows": row_count, "sheet_names": wb.sheetnames})
        wb.close()
    for path in sorted((archive_root / "study_phase2/ratings").glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=False)
        files.append({"file": path.relative_to(archive_root).as_posix(), "bytes": path.stat().st_size, "data_rows": None, "sheet_names": wb.sheetnames})
        wb.close()
    report = {"schema_version": "archive-excel-validation-1.0.0", "status": "passed" if not issues else "failed", "issues": issues, "files": files}
    write_json(archive_root / "validation/excel_integrity.json", report)
    if issues:
        raise RuntimeError(f"archive Excel validation failed: {issues}")
    return report


def _copy_sanitized_json(source: Path, target: Path, root: Path) -> None:
    value = json.loads(source.read_text(encoding="utf-8"))
    write_json(target, sanitize_value(value, root))


def build_redacted_reports(root: Path, archive_root: Path) -> None:
    mappings = {
        "outputs/study_phase2/manifest/study_manifest.json": "study_phase2/manifest/study_manifest_redacted.json",
        "outputs/study_phase2/manifest/study_manifest_at_freeze_v2_500usd.json": "study_phase2/manifest/study_manifest_at_freeze_v2_500usd_redacted.json",
        "outputs/study_phase2/questions/study_owner_pre_freeze_approval.json": "study_phase2/questions/study_owner_pre_freeze_approval_redacted.json",
        "outputs/study_phase2/qa/phase2_validation.json": "study_phase2/qa/phase2_validation_redacted.json",
        "outputs/study_phase2/qa/excel_integrity.json": "study_phase2/qa/excel_integrity_redacted.json",
        "outputs/study_phase2/reports/technical_completion_report.json": "study_phase2/reports/technical_completion_report_redacted.json",
    }
    for source, target in mappings.items():
        _copy_sanitized_json(root / source, archive_root / target, root)


def _is_core_file(relative: str) -> bool:
    return relative in ROOT_FILES or relative.startswith(("src/", "scripts/", "tests/", "db/", "docs/"))


def _iter_project_files(root: Path) -> Iterable[Path]:
    """Yield files while representing bulky generated directories by rules."""

    for directory, dirnames, filenames in os.walk(root):
        dirnames[:] = sorted(name for name in dirnames if name not in PRUNED_LOCAL_DIRECTORIES)
        base = Path(directory)
        for filename in sorted(filenames):
            path = base / filename
            if path.is_file():
                yield path


def archive_allowlist(root: Path) -> list[str]:
    included: set[str] = set()
    for path in _iter_project_files(root):
        relative = path.relative_to(root).as_posix()
        if _is_core_file(relative) or relative in SAFE_OUTPUT_FILES or relative.startswith(f"{ARCHIVE_ROOT_NAME}/"):
            included.add(relative)
    missing = sorted(path for path in ROOT_FILES | SAFE_OUTPUT_FILES if not (root / path).is_file())
    if missing:
        raise RuntimeError(f"archive allowlist references missing files: {missing}")
    return sorted(included)


def exclusion_reason(relative: str) -> tuple[str, str]:
    if relative == ".env.retrieval" or (Path(relative).name.startswith(".env") and relative != ".env.example"):
        return "secret_environment", "Real environment or database secret; never distributed."
    if relative.startswith("source_pdfs/") or relative.lower().endswith(".pdf"):
        return "rights_restricted_source", "Original guideline or medicinal product PDF; rights not cleared for redistribution."
    if relative.startswith(("outputs/knowledge_corpus/canonical/", "data/canonical/")):
        return "rights_restricted_fulltext", "Complete extracted canonical corpus; redistribution rights not established."
    if "/embeddings/" in relative or "/query_embeddings/" in relative or "embedding_input" in relative:
        return "vector_or_embedding", "Corpus/query embedding vector or embedding input; excluded by archive policy."
    if relative.startswith(("outputs/knowledge_corpus/checkpoints/", "data/extraction_runs/")):
        return "checkpoint_or_raw_extraction", "Raw extraction/checkpoint material may contain complete source text or provider payloads."
    if relative in {
        "outputs/study_phase2/results/api_attempts.jsonl",
        "outputs/study_phase2/results/api_attempts.csv",
        "outputs/study_phase2/results/study_results.jsonl",
        "outputs/study_phase2/results/study_results.csv",
    }:
        return "raw_operational_study_data", "Unredacted operational identifiers/headers or full retrieval/tool payloads; redacted derivative is archived."
    if relative.startswith("outputs/study_phase2/results/"):
        return "raw_or_redundant_study_data", "Raw/redundant study ledger; canonical redacted derivative or aggregate is archived."
    if relative.startswith(("outputs/study_phase2/excel/", "outputs/study_phase2/ratings/")):
        return "unredacted_office_export", "Historical workbook retained locally; separately redacted workbook is archived."
    if relative.startswith((".venv/", ".pytest_cache/", ".ruff_cache/")) or "/__pycache__/" in relative:
        return "generated_environment_or_cache", "Generated environment/cache; reproduced from lockfile."
    if relative.startswith(".git/"):
        return "git_internal", "Git internal metadata is outside the repository content manifest."
    if relative.startswith("outputs/"):
        return "non_allowlisted_output", "Generated output not required or not cleared for redistribution."
    if relative.startswith("data/"):
        return "non_allowlisted_data", "Local generated/source-derived data not cleared for redistribution."
    return "not_required", "Not part of the reviewed reproducibility allowlist."


def build_repository_manifest(root: Path, archive_root: Path) -> dict[str, Any]:
    allowlist_path = archive_root / "repository_allowlist.txt"
    decisions_path = archive_root / "repository_file_decisions.json"
    decisions_csv_path = archive_root / "repository_file_decisions.csv"
    checksums_path = archive_root / "ARCHIVE_SHA256SUMS"
    # These files are generated at the end and are deliberately part of the allowlist.
    for path in (allowlist_path, decisions_path, decisions_csv_path, checksums_path):
        path.parent.mkdir(parents=True, exist_ok=True)
        if not path.exists():
            path.touch()
    included = archive_allowlist(root)
    allowlist_path.write_text("\n".join(included) + "\n", encoding="utf-8")
    included = archive_allowlist(root)
    decisions: list[dict[str, Any]] = []
    self_metadata = {
        decisions_path.relative_to(root).as_posix(),
        decisions_csv_path.relative_to(root).as_posix(),
        checksums_path.relative_to(root).as_posix(),
    }
    for path in _iter_project_files(root):
        relative = path.relative_to(root).as_posix()
        if relative in included:
            decisions.append(
                {
                    "path": relative,
                    "decision": "include",
                    "category": "reviewed_reproducibility_artifact",
                    "reason": "Explicitly selected by the archive allowlist and subject to repository validation.",
                    "bytes": None if relative in self_metadata else path.stat().st_size,
                    "sha256": None if relative in self_metadata else sha256_file(path),
                }
            )
        else:
            category, reason = exclusion_reason(relative)
            decisions.append(
                {
                    "path": relative,
                    "decision": "exclude",
                    "category": category,
                    "reason": reason,
                    "bytes": path.stat().st_size,
                    "sha256": None,
                }
            )
    payload = {
        "schema_version": "repository-file-decisions-1.0.0",
        "scope_note": ".git internals, virtual environments and cache trees are covered by excluded_path_rules rather than enumerated file-by-file.",
        "excluded_path_rules": [
            {
                "pattern": f"**/{name}/**" if name == "__pycache__" else f"{name}/**",
                "decision": "exclude",
                "category": category,
                "reason": reason,
            }
            for name, (category, reason) in sorted(PRUNED_LOCAL_DIRECTORIES.items())
        ],
        "included_count": sum(item["decision"] == "include" for item in decisions),
        "excluded_count": sum(item["decision"] == "exclude" for item in decisions),
        "decisions": decisions,
    }
    write_json(decisions_path, payload)
    write_csv(decisions_csv_path, decisions, fields=("path", "decision", "category", "reason", "bytes", "sha256"))
    included = archive_allowlist(root)
    allowlist_path.write_text("\n".join(included) + "\n", encoding="utf-8")
    checksum_lines = [
        f"{sha256_file(root / relative)}  {relative}"
        for relative in included
        if relative != checksums_path.relative_to(root).as_posix()
    ]
    checksums_path.write_text("\n".join(checksum_lines) + "\n", encoding="utf-8")
    return payload


def _scan_text(path: Path) -> list[str]:
    if path.suffix.lower() in {".xlsx", ".gz", ".png", ".jpg", ".jpeg", ".gif", ".zip"}:
        return []
    try:
        text = path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return []
    issues = [name for name, pattern in SECRET_PATTERNS.items() if pattern.search(text)]
    if ABSOLUTE_USER_PATH.search(text):
        issues.append("absolute_user_path")
    return issues


def _scan_workbook(path: Path) -> list[str]:
    issues: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if not isinstance(value, str):
                    continue
                for name, pattern in SECRET_PATTERNS.items():
                    if pattern.search(value):
                        issues.append(f"{name}:{ws.title}")
                if ABSOLUTE_USER_PATH.search(value):
                    issues.append(f"absolute_user_path:{ws.title}")
    wb.close()
    return sorted(set(issues))


def build_security_report(root: Path, archive_root: Path) -> dict[str, Any]:
    """Scan the complete project scope without ever recording matched values."""

    included = set(archive_allowlist(root))
    blocking: list[dict[str, Any]] = []
    excluded_findings: list[dict[str, Any]] = []
    temporary_office_files: list[str] = []
    files_over_50_mib: list[dict[str, Any]] = []
    files_over_100_mib: list[dict[str, Any]] = []
    scanned_files = 0
    for path in _iter_project_files(root):
        relative = path.relative_to(root).as_posix()
        scanned_files += 1
        size = path.stat().st_size
        if size > 50 * 1024 * 1024:
            files_over_50_mib.append({"path": relative, "bytes": size, "included": relative in included})
        if size > 100 * 1024 * 1024:
            files_over_100_mib.append({"path": relative, "bytes": size, "included": relative in included})
        if path.name.startswith(("~$", ".~lock.")):
            temporary_office_files.append(relative)
        findings = _scan_workbook(path) if path.suffix.lower() == ".xlsx" else _scan_text(path)
        if path.name.startswith(".env") and path.name != ".env.example":
            try:
                nonempty_secret_names = []
                for line in path.read_text(encoding="utf-8").splitlines():
                    stripped = line.strip()
                    if not stripped or stripped.startswith("#") or "=" not in stripped:
                        continue
                    name, value = stripped.split("=", 1)
                    if value.strip() and any(term in name.upper() for term in ("KEY", "TOKEN", "PASSWORD", "SECRET", "AUTH")):
                        nonempty_secret_names.append(name.strip())
                if nonempty_secret_names:
                    findings.append("nonempty_secret_environment_value")
            except UnicodeDecodeError:
                findings.append("unreadable_secret_environment_file")
        for finding in sorted(set(findings)):
            entry = {"path": relative, "finding_class": finding, "included": relative in included}
            is_absolute_path_finding = finding.startswith("absolute_user_path")
            if relative in included or (not is_absolute_path_finding and path.name not in KNOWN_LOCAL_SECRET_FILES):
                blocking.append(entry)
            else:
                excluded_findings.append(entry)

    approval = json.loads(
        (root / "outputs/study_phase2/questions/study_owner_pre_freeze_approval.json").read_text(encoding="utf-8")
    )
    questions = read_jsonl(root / "outputs/study_phase2/questions/study_questions_frozen.jsonl")
    question_checks = {
        "question_count": len(questions),
        "all_synthetic": all("synthetic" in str(row.get("authoring_method", "")) for row in questions),
        "patient_identifier_fields_present": any(
            key in row
            for row in questions
            for key in ("patient_id", "patient_name", "date_of_birth", "medical_record_number")
        ),
        "reviewer_name_recorded": approval.get("reviewer_name_recorded"),
        "signature_recorded": approval.get("signature_recorded"),
        "independent_clinical_question_validation_claimed": approval.get(
            "independent_clinical_question_validation_claimed"
        ),
    }
    if len(questions) != 100 or not question_checks["all_synthetic"] or question_checks["patient_identifier_fields_present"]:
        blocking.append({"path": "outputs/study_phase2/questions/study_questions_frozen.jsonl", "finding_class": "question_privacy_or_cardinality", "included": True})
    if question_checks["reviewer_name_recorded"] or question_checks["signature_recorded"]:
        blocking.append({"path": "outputs/study_phase2/questions/study_owner_pre_freeze_approval.json", "finding_class": "reviewer_identity_or_signature", "included": False})
    if any(item["included"] for item in files_over_50_mib):
        blocking.append({"path": "archive/repository_allowlist.txt", "finding_class": "included_file_over_50_mib", "included": True})
    report = {
        "schema_version": "repository-security-scan-1.0.0",
        "scope": "all project files except .git internals, virtual environments and cache trees",
        "status": "passed" if not blocking else "failed",
        "scanned_file_count": scanned_files,
        "blocking_findings": blocking,
        "expected_excluded_findings": excluded_findings,
        "temporary_office_files": temporary_office_files,
        "files_over_50_mib": files_over_50_mib,
        "files_over_100_mib": files_over_100_mib,
        "question_and_reviewer_checks": question_checks,
        "operational_data_policy": {
            "raw_api_attempts_included": "outputs/study_phase2/results/api_attempts.jsonl" in included,
            "redacted_attempts_included": "archive/study_phase2/results/api_attempts_redacted.jsonl" in included,
            "source_pdfs_included": any(path.startswith("source_pdfs/") for path in included),
            "canonical_fulltext_included": any(path.startswith("outputs/knowledge_corpus/canonical/") for path in included),
            "embedding_vectors_included": any("/full/ecp-" in path for path in included),
        },
        "note": "Only finding classes and file paths are recorded; matched values are never persisted.",
    }
    write_json(archive_root / "validation/security_scan.json", report)
    if blocking:
        raise RuntimeError(f"repository security scan found {len(blocking)} blocking finding(s)")
    return report


def build_hash_integrity_report(root: Path, archive_root: Path) -> dict[str, Any]:
    """Verify immutable corpus and study manifests without rewriting sources."""

    snapshot_path = root / f"outputs/knowledge_corpus/manifests/corpus_snapshots/{SNAPSHOT_ID}.json"
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    groups: list[tuple[str, list[dict[str, Any]]]] = [
        (
            "source_pdfs",
            [
                {"path": item["relative_path"], "sha256": item["source_sha256"], "bytes": item["file_size_bytes"]}
                for item in snapshot["source_pdfs"]
            ],
        ),
        (
            "canonical_files",
            [
                {"path": item["relative_path"], "sha256": item["sha256"], "bytes": item["size_bytes"]}
                for item in snapshot["canonical_files"]
            ],
        ),
        (
            "snapshot_provenance_artifacts",
            [
                {"path": item["relative_path"], "sha256": item["sha256"], "bytes": item["size_bytes"]}
                for item in snapshot["artifact_integrity"]
            ],
        ),
    ]
    for name, filename in (
        ("study_owner_freeze", "artifact_hashes_study_owner_freeze_v2_500usd.json"),
        ("technical_completion", "artifact_hashes_technical_complete.json"),
    ):
        manifest = json.loads((root / f"outputs/study_phase2/manifest/{filename}").read_text(encoding="utf-8"))
        groups.append((name, list(manifest["artifacts"])))

    summaries: list[dict[str, Any]] = []
    mismatches: list[dict[str, Any]] = []
    for group_name, artifacts in groups:
        matched = 0
        for artifact in artifacts:
            relative = str(artifact["path"])
            path = root / relative
            issue = None
            if not path.is_file():
                issue = "missing"
            elif artifact.get("bytes") is not None and path.stat().st_size != int(artifact["bytes"]):
                issue = "size_mismatch"
            elif sha256_file(path) != artifact["sha256"]:
                issue = "sha256_mismatch"
            else:
                matched += 1
            if issue:
                mismatches.append({"group": group_name, "path": relative, "issue": issue})
        summaries.append({"group": group_name, "expected": len(artifacts), "matched": matched})
    report = {
        "schema_version": "archive-hash-integrity-1.0.0",
        "status": "passed" if not mismatches else "failed",
        "corpus_snapshot_id": SNAPSHOT_ID,
        "groups": summaries,
        "mismatches": mismatches,
        "technical_completion_manifest_sha256": sha256_file(
            root / "outputs/study_phase2/manifest/artifact_hashes_technical_complete.json"
        ),
    }
    write_json(archive_root / "validation/hash_integrity.json", report)
    if mismatches:
        raise RuntimeError(f"immutable artifact integrity failed for {len(mismatches)} file(s)")
    return report


def validate_archive(
    root: Path,
    *,
    require_allowlist: bool = True,
    write_report: bool = True,
) -> dict[str, Any]:
    root = root.resolve()
    archive_root = root / ARCHIVE_ROOT_NAME
    allowlist_file = archive_root / "repository_allowlist.txt"
    if require_allowlist and not allowlist_file.is_file():
        raise RuntimeError("archive allowlist is missing")
    included = [line.strip() for line in allowlist_file.read_text(encoding="utf-8").splitlines() if line.strip()]
    issues: list[dict[str, Any]] = []
    total_bytes = 0
    for relative in included:
        path = root / relative
        if not path.is_file():
            issues.append({"path": relative, "issue": "missing"})
            continue
        size = path.stat().st_size
        total_bytes += size
        if size > 50 * 1024 * 1024:
            issues.append({"path": relative, "issue": "over_50_mib", "bytes": size})
        if size > 100 * 1024 * 1024:
            issues.append({"path": relative, "issue": "over_100_mib", "bytes": size})
        if path.suffix.lower() == ".pdf":
            issues.append({"path": relative, "issue": "pdf_forbidden"})
        if relative != ".env.example" and Path(relative).name.startswith(".env"):
            issues.append({"path": relative, "issue": "environment_file_forbidden"})
        if any(relative.startswith(prefix) for prefix in FORBIDDEN_TRACKED_PREFIXES):
            issues.append({"path": relative, "issue": "forbidden_path"})
        for issue in _scan_workbook(path) if path.suffix.lower() == ".xlsx" else _scan_text(path):
            issues.append({"path": relative, "issue": issue})
    if len(included) != len(set(included)):
        issues.append({"path": str(allowlist_file), "issue": "duplicate_allowlist_path"})
    results_path = archive_root / "study_phase2/results/study_results_redacted.jsonl"
    if results_path.is_file():
        rows = read_jsonl(results_path)
        if len(rows) != 800 or len({row.get("run_id") for row in rows}) != 800:
            issues.append({"path": str(results_path), "issue": "study_result_cardinality"})
        expected = {
            ("gpt55_medium", "WEB"): 200,
            ("gpt55_medium", "RAG"): 200,
            ("gpt56_sol_high", "WEB"): 200,
            ("gpt56_sol_high", "RAG"): 200,
        }
        observed = {
            key: sum(row.get("model_config_id") == key[0] and row.get("system_arm") == key[1] for row in rows)
            for key in expected
        }
        if observed != expected:
            issues.append({"path": str(results_path), "issue": "study_cell_distribution", "observed": observed})
    source_manifest = json.loads((archive_root / "corpus/source_manifest.json").read_text(encoding="utf-8"))
    if source_manifest.get("source_count") != 12 or source_manifest.get("page_count") != 2060:
        issues.append({"path": "archive/corpus/source_manifest.json", "issue": "source_manifest_counts"})
    security_path = archive_root / "validation/security_scan.json"
    if not security_path.is_file() or json.loads(security_path.read_text(encoding="utf-8")).get("status") != "passed":
        issues.append({"path": "archive/validation/security_scan.json", "issue": "security_scan_missing_or_failed"})
    hash_path = archive_root / "validation/hash_integrity.json"
    if not hash_path.is_file() or json.loads(hash_path.read_text(encoding="utf-8")).get("status") != "passed":
        issues.append({"path": "archive/validation/hash_integrity.json", "issue": "hash_integrity_missing_or_failed"})
    excel = validate_excel_exports(archive_root)
    if excel["status"] != "passed":
        issues.append({"path": "archive/validation/excel_integrity.json", "issue": "excel_validation"})
    build_report_path = archive_root / "archive_build_report.json"
    build_timestamp = None
    if build_report_path.is_file():
        build_timestamp = json.loads(build_report_path.read_text(encoding="utf-8")).get("built_at_utc")
    report = {
        "schema_version": "github-archive-validation-1.0.0",
        "archive_build_timestamp_utc": build_timestamp,
        "status": "passed" if not issues else "failed",
        "included_file_count": len(included),
        "included_bytes": total_bytes,
        "included_mib": round(total_bytes / (1024 * 1024), 3),
        "largest_file_bytes": max(((root / item).stat().st_size for item in included), default=0),
        "issues": issues,
        "assertions": {
            "no_source_pdfs": not any(item.lower().endswith(".pdf") for item in included),
            "no_canonical_fulltext": not any(item.startswith("outputs/knowledge_corpus/canonical/") for item in included),
            "no_embedding_vectors": not any("/full/ecp-" in item or "/query_embeddings/" in item for item in included),
            "no_raw_api_attempts": "outputs/study_phase2/results/api_attempts.jsonl" not in included,
            "no_git_lfs_required": all((root / item).stat().st_size <= 50 * 1024 * 1024 for item in included),
            "study_results": 800,
            "corpus_snapshot_id": SNAPSHOT_ID,
        },
    }
    if write_report:
        write_json(archive_root / "validation/archive_validation.json", report)
    if issues:
        raise RuntimeError(f"archive validation failed with {len(issues)} issue(s)")
    return report


def build_archive(root: Path) -> dict[str, Any]:
    root = root.resolve()
    archive_root = root / ARCHIVE_ROOT_NAME
    archive_root.mkdir(parents=True, exist_ok=True)
    build_source_manifest(root, archive_root)
    results, attempts = build_redacted_machine_results(root, archive_root)
    build_redacted_reports(root, archive_root)
    excel_report = build_excel_exports(root, archive_root, results, attempts)
    hash_report = build_hash_integrity_report(root, archive_root)
    security_report = build_security_report(root, archive_root)
    build_report = {
        "schema_version": ARCHIVE_SCHEMA_VERSION,
        "built_at_utc": utc_now(),
        "corpus_snapshot_id": SNAPSHOT_ID,
        "retrieval_units": 4469,
        "existing_corpus_embeddings_excluded": 4469,
        "study_question_count": 100,
        "study_result_count": len(results),
        "api_attempt_count": len(attempts),
        "excel_status": excel_report["status"],
        "security_status": security_report["status"],
        "hash_integrity_status": hash_report["status"],
        "historical_artifacts_modified": False,
        "external_api_calls": 0,
    }
    write_json(archive_root / "archive_build_report.json", build_report)
    build_repository_manifest(root, archive_root)
    validation = validate_archive(root)
    build_repository_manifest(root, archive_root)
    validation = validate_archive(root, write_report=False)
    return {"build": build_report, "validation": validation}


__all__ = [
    "ARCHIVE_SCHEMA_VERSION",
    "SAFE_OUTPUT_FILES",
    "SNAPSHOT_ID",
    "archive_allowlist",
    "build_archive",
    "validate_archive",
]
