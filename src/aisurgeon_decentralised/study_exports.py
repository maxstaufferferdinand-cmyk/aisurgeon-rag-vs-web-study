"""Deterministic JSONL/CSV-derived Excel and human-review exports."""

from __future__ import annotations

import csv
import json
import random
import re
from collections.abc import Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .study_phase2 import (
    RANDOMIZATION_SEED,
    PlannedStudyCell,
    StudyQuestion,
    build_randomization_manifest,
    read_jsonl,
    validate_question_set,
    write_jsonl_atomic,
)

WEB_FILL = "F4B183"
RAG_FILL = "9DC3E6"
GPT55_FILL = "A9D18E"
GPT56_FILL = "B4A7D6"
RUN1_FILL = "5B9BD5"
RUN2_FILL = "D9EAF7"
HEADER_FILL = "1F4E78"

RESULT_COLUMNS = [
    "Frage_ID",
    "Corpus_Abdeckung",
    "Modell",
    "Reasoning",
    "System_WEB_oder_RAG",
    "Run",
    "Frage",
    "Antwortstatus",
    "Validierte_Antwort",
    "Rohantwort",
    "Quellen",
    "Input_Tokens",
    "Cached_Input_Tokens",
    "Output_Tokens",
    "Reasoning_Tokens",
    "Total_Tokens",
    "Web_Search_Aufrufe",
    "Retrievalzeit_ms",
    "API_Zeit_ms",
    "End_to_End_Zeit_ms",
    "Geschaetzte_Kosten_USD",
    "Abgeglichene_Kosten_USD",
    "Validatorstatus",
    "Fehler_oder_Retry",
    "Run_ID",
    "Geplante_Reihenfolge",
    "Tatsaechliche_Reihenfolge",
    "Randomisierungsblock",
    "Status",
    "Requested_Model",
    "Returned_Model",
    "Question_Hash",
    "Corpus_Snapshot_ID",
    "Response_ID",
    "Request_ID",
    "Client_Request_ID",
    "Evidence_Allowlist",
    "Web_Quellen_konsultiert",
    "Retry_Count",
    "Embedding_Tokens",
    "Embedding_Cache_Hit",
    "OpenAI_Processing_ms",
    "Time_to_First_Token_ms",
    "Web_Search_Zeit_ms",
    "DB_Zeit_ms",
    "Relationsexpansion_ms",
    "Standardisierte_Uncached_Kosten_USD",
    "Kostenabgleich_Status",
    "Start_UTC",
    "Ende_UTC",
    "Protocol_Deviation_IDs",
    "Cache_Write_Tokens",
    "Search_Content_Tokens",
    "Web_Quellen_zitiert",
    "Web_Search_Aktionen",
    "Lokale_Ressourcen",
    "Rate_Limit_Headers",
    "HTTP_Status",
    "Service_Tier_Requested",
    "Service_Tier_Used",
    "Prompt_Hashes",
    "Response_Schema_Hash",
    "Retrieval_Config_Hash",
    "Web_Config_Hash",
    "Preisversion",
    "Validator_Issue_Codes",
    "Evidence_Package_ID",
    "Query_Normalisierung_ms",
    "Exakte_Suche_ms",
    "FTS_ms",
    "Trigramm_ms",
    "Vektorsuche_ms",
    "RRF_ms",
    "Evidence_Package_ms",
]


def _json_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _display_model(config_id: str) -> str:
    return {
        "gpt55_medium": "GPT-5.5",
        "gpt56_sol_high": "GPT-5.6 Sol",
    }[config_id]


def _planned_result_row(
    cell: PlannedStudyCell,
    question: StudyQuestion,
    result: dict[str, Any] | None,
    attempts: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    validated = (result or {}).get("validated_system_answer") or {}
    raw = (result or {}).get("raw_model_answer")
    token = (result or {}).get("token_usage") or {}
    timing = (result or {}).get("timing_ms") or {}
    cost = (result or {}).get("cost") or {}
    retrieval = (result or {}).get("retrieval") or {}
    sources = (result or {}).get("backend_rendered_sources") or ()
    labels = [source.get("label") or source.get("url") for source in sources]
    attempts_for_run = [row for row in attempts if row.get("run_id") == cell.run_id]
    final_attempt = attempts_for_run[-1] if attempts_for_run else {}
    return dict(
        zip(
            RESULT_COLUMNS,
            [
                question.question_id,
                question.coverage_stratum,
                _display_model(cell.model_config_id),
                cell.reasoning_effort,
                cell.system_arm,
                cell.repetition,
                question.question_text,
                validated.get("answer_status", ""),
                validated.get("answer_text", ""),
                _json_cell(raw),
                "\n".join(filter(None, labels)),
                token.get("input_tokens", ""),
                token.get("cached_input_tokens", ""),
                token.get("output_tokens", ""),
                token.get("reasoning_tokens", ""),
                token.get("total_tokens", ""),
                final_attempt.get("web_search_tool_calls", ""),
                retrieval.get("retrieval_time_ms", ""),
                timing.get("api_wall", ""),
                timing.get("end_to_end", ""),
                cost.get("total_estimated_cost_usd", ""),
                cost.get("reconciled_cost_usd", ""),
                (result or {}).get("validator_status", ""),
                (result or {}).get("error_code", "")
                or (
                    f"Retries: {(result or {}).get('retry_count')}"
                    if (result or {}).get("retry_count")
                    else ""
                ),
                cell.run_id,
                cell.planned_order,
                (result or {}).get("actual_order", ""),
                cell.randomization_block,
                (result or {}).get("status", cell.status),
                cell.requested_model,
                (result or {}).get("returned_model", ""),
                question.question_hash,
                (result or {}).get("corpus_snapshot_id", "cs-f61b3d4e90089c1b890c23cb"),
                final_attempt.get("response_id", ""),
                final_attempt.get("x_request_id", ""),
                final_attempt.get("client_request_id", ""),
                _json_cell((result or {}).get("evidence_allowlist", ())),
                _json_cell((result or {}).get("web_sources_consulted", ())),
                (result or {}).get("retry_count", 0),
                token.get("query_embedding_tokens", ""),
                retrieval.get("embedding_cache_hit", ""),
                timing.get("openai_processing", ""),
                timing.get("time_to_first_token", ""),
                timing.get("web_search", ""),
                timing.get("database", ""),
                timing.get("relation_expansion", ""),
                cost.get("standardized_uncached_cost_usd", ""),
                cost.get("cost_reconciliation_status", ""),
                (result or {}).get("started_at_utc", ""),
                (result or {}).get("finished_at_utc", ""),
                _json_cell((result or {}).get("protocol_deviation_ids", ())),
                token.get("cache_write_tokens", ""),
                token.get("search_content_tokens", ""),
                _json_cell((result or {}).get("web_sources_cited", ())),
                _json_cell((result or {}).get("web_search_actions", ())),
                _json_cell((result or {}).get("local_resources", {})),
                _json_cell(final_attempt.get("rate_limit_headers", {})),
                final_attempt.get("http_status", ""),
                final_attempt.get("service_tier_requested", ""),
                final_attempt.get("service_tier_used", ""),
                _json_cell((result or {}).get("prompt_hashes", {})),
                (result or {}).get("response_schema_hash", ""),
                (result or {}).get("retrieval_config_hash", ""),
                (result or {}).get("web_config_hash", ""),
                cost.get("price_version", final_attempt.get("price_version", "")),
                _json_cell((result or {}).get("validator_issue_codes", ())),
                (result or {}).get("evidence_package_id", ""),
                timing.get("query_normalization", ""),
                timing.get("exact_search", ""),
                timing.get("fts", ""),
                timing.get("trigram", ""),
                timing.get("vector", ""),
                timing.get("rrf", ""),
                timing.get("evidence_package", ""),
            ],
            strict=True,
        )
    )


def build_planned_rows(
    *, root: Path, questions: Sequence[StudyQuestion]
) -> tuple[list[dict[str, Any]], tuple[PlannedStudyCell, ...]]:
    cells = build_randomization_manifest(questions, frozen=False)
    result_rows = read_jsonl(root / "outputs/study_phase2/results/study_results.jsonl")
    attempts = read_jsonl(root / "outputs/study_phase2/results/api_attempts.jsonl")
    by_result = {row["run_id"]: row for row in result_rows}
    by_question = {row.question_id: row for row in questions}
    rows = [
        _planned_result_row(
            cell, by_question[cell.question_id], by_result.get(cell.run_id), attempts
        )
        for cell in sorted(cells, key=lambda row: row.planned_order)
    ]
    return rows, cells


def export_planned_results(
    *, root: Path, questions: Sequence[StudyQuestion]
) -> dict[str, Path]:
    rows, _ = build_planned_rows(root=root, questions=questions)
    directory = root / "outputs/study_phase2/results"
    directory.mkdir(parents=True, exist_ok=True)
    jsonl = directory / "planned_results.jsonl"
    csv_path = directory / "planned_results.csv"
    write_jsonl_atomic(jsonl, rows)
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=RESULT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    return {"jsonl": jsonl, "csv": csv_path}


def _style_header(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _add_table(sheet: Any, *, name: str) -> None:
    if sheet.max_row < 2 or sheet.max_column < 1:
        return
    table = Table(
        displayName=name,
        ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _set_widths(sheet: Any, *, result_sheet: bool = False) -> None:
    for index, cell in enumerate(sheet[1], start=1):
        header = str(cell.value or "")
        width = max(12, min(28, len(header) + 3))
        if header in {"Frage", "Validierte_Antwort", "Rohantwort", "Quellen"}:
            width = 48 if header != "Rohantwort" else 38
        sheet.column_dimensions[get_column_letter(index)].width = width
    if result_sheet:
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _append_dict_rows(
    sheet: Any, rows: Sequence[dict[str, Any]], columns: Sequence[str]
) -> None:
    sheet.append(list(columns))
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _style_results(sheet: Any) -> None:
    header = {cell.value: cell.column for cell in sheet[1]}
    for row in range(2, sheet.max_row + 1):
        arm = sheet.cell(row, header["System_WEB_oder_RAG"])
        model = sheet.cell(row, header["Modell"])
        repetition = sheet.cell(row, header["Run"])
        arm.fill = PatternFill(
            "solid", fgColor=WEB_FILL if arm.value == "WEB" else RAG_FILL
        )
        model.fill = PatternFill(
            "solid", fgColor=GPT55_FILL if model.value == "GPT-5.5" else GPT56_FILL
        )
        repetition.fill = PatternFill(
            "solid", fgColor=RUN1_FILL if repetition.value == "1_primary" else RUN2_FILL
        )
        if repetition.value == "1_primary":
            repetition.font = Font(bold=True, color="FFFFFF")
        for name in ("Geschaetzte_Kosten_USD", "Abgeglichene_Kosten_USD"):
            sheet.cell(row, header[name]).number_format = "0.000000"


def build_question_review_workbook(
    *, questions: Sequence[StudyQuestion], path: Path
) -> None:
    validate_question_set(questions, require_human_freeze=False)
    workbook = Workbook()
    guide = workbook.active
    guide.title = "00_ANLEITUNG"
    instructions = [
        [
            "Zweck",
            "Unabhängige klinische Freigabe des synthetischen 100-Fragen-Testsets vor API-Hauptlauf.",
        ],
        [
            "Reviewer A/B",
            "Unabhängig Coverage, Pflichtclaims, kritische Fehler und Goldquellen/RAG-Abstention prüfen.",
        ],
        [
            "Adjudikation",
            "Divergenzen nach beiden Erstbewertungen dokumentiert auflösen.",
        ],
        [
            "Freigabe",
            "Nur approved + vier TRUE-Bestätigungen + Freeze-Zeitstempel dürfen nach study_questions_frozen.jsonl exportiert werden.",
        ],
        [
            "Coverage",
            "not_covered bedeutet nur: im lokalen Snapshot nicht ausreichend abgedeckt; nicht allgemein medizinisch unbeantwortbar.",
        ],
        [
            "Zulässige Werte",
            "Die Reviewfelder bleiben beim study_owner_pre_freeze_approval leer; keine Reviewer, Kommentare oder Signaturen erfinden.",
        ],
        [
            "HCC",
            "Die HCC/BCC-Quelle ist consultation_draft und muss so behandelt werden.",
        ],
        [
            "Gold",
            "Study-owner-freigegebene Drafts sind kein unabhängig klinisch validierter Goldstandard; spätere verblindete Antwortreviews bleiben erforderlich.",
        ],
    ]
    for row in instructions:
        guide.append(row)
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 110
    guide.freeze_panes = "A1"
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    guide["A1"].font = Font(bold=True)

    sheet = workbook.create_sheet("01_FRAGEN_GOLD")
    columns = [
        "question_id",
        "question_text",
        "coverage_stratum_provisional",
        "clinical_domain",
        "question_type",
        "difficulty",
        "required_claims_provisional",
        "critical_omissions_provisional",
        "forbidden_or_harmful_claims_provisional",
        "expected_formal_item_ids",
        "expected_retrieval_unit_ids",
        "expected_source_documents",
        "expected_pages",
        "expected_active_substance_ids",
        "expected_product_ids",
        "expected_relation_types",
        "source_status_notes",
        "Reviewer_A_Name",
        "Reviewer_A_Coverage",
        "Reviewer_A_Pflichtclaims_bestaetigt",
        "Reviewer_A_kritische_Fehler_bestaetigt",
        "Reviewer_A_Goldquellen_oder_Abstention_bestaetigt",
        "Reviewer_A_Status",
        "Reviewer_A_Kommentar",
        "Reviewer_B_Name",
        "Reviewer_B_Coverage",
        "Reviewer_B_Pflichtclaims_bestaetigt",
        "Reviewer_B_kritische_Fehler_bestaetigt",
        "Reviewer_B_Goldquellen_oder_Abstention_bestaetigt",
        "Reviewer_B_Status",
        "Reviewer_B_Kommentar",
        "Adjudikator_Name",
        "Adjudizierte_Coverage",
        "Adjudizierte_Pflichtclaims",
        "Adjudizierte_kritische_Fehler",
        "Adjudizierte_Goldquellen_oder_Abstention",
        "Adjudikationsstatus",
        "Adjudikationskommentar",
        "Freeze_Timestamp_UTC",
    ]
    rows = []
    for question in questions:
        payload = question.model_dump(mode="json")
        row = {
            "question_id": question.question_id,
            "question_text": question.question_text,
            "coverage_stratum_provisional": question.coverage_stratum,
            "clinical_domain": question.clinical_domain,
            "question_type": question.question_type,
            "difficulty": question.difficulty,
            "required_claims_provisional": _json_cell(payload["required_claims"]),
            "critical_omissions_provisional": _json_cell(payload["critical_omissions"]),
            "forbidden_or_harmful_claims_provisional": _json_cell(
                payload["forbidden_or_harmful_claims"]
            ),
            "expected_formal_item_ids": _json_cell(payload["expected_formal_item_ids"]),
            "expected_retrieval_unit_ids": _json_cell(
                payload["expected_retrieval_unit_ids"]
            ),
            "expected_source_documents": _json_cell(
                payload["expected_source_documents"]
            ),
            "expected_pages": _json_cell(payload["expected_pages"]),
            "expected_active_substance_ids": _json_cell(
                payload["expected_active_substance_ids"]
            ),
            "expected_product_ids": _json_cell(payload["expected_product_ids"]),
            "expected_relation_types": _json_cell(payload["expected_relation_types"]),
            "source_status_notes": _json_cell(payload["source_status_notes"]),
        }
        rows.append(row)
    _append_dict_rows(sheet, rows, columns)
    _add_table(sheet, name="QuestionFreezeReview")
    _set_widths(sheet, result_sheet=True)
    sheet.column_dimensions["B"].width = 60
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _blind_rows(
    rows: Sequence[dict[str, Any]], *, include_sources: bool
) -> list[dict[str, Any]]:
    primary = [row for row in rows if row["Run"] == "1_primary"]
    rng = random.Random(RANDOMIZATION_SEED)
    rng.shuffle(primary)
    output = []
    for index, row in enumerate(primary, start=1):
        blinded = f"BR-{index:04d}"
        answer = str(row["Validierte_Antwort"] or "")
        if not include_sources:
            answer = re.sub(r"https?://\S+", "[QUELLE ENTFERNT]", answer)
            answer = re.sub(
                r"\bru-[0-9a-f]{8,}\b",
                "[LOKALE QUELLEN-ID ENTFERNT]",
                answer,
                flags=re.IGNORECASE,
            )
        base = {
            "blinded_response_id": blinded,
            "question_id": row["Frage_ID"],
            "question": row["Frage"],
            "answer_status": row["Antwortstatus"],
            "answer": answer,
        }
        if include_sources:
            base.update(
                {
                    "model": row["Modell"],
                    "system_arm": row["System_WEB_oder_RAG"],
                    "sources": row["Quellen"],
                    "run_id": row["Run_ID"],
                }
            )
        else:
            base.update(
                {
                    "reviewer_a_name": "",
                    "reviewer_a_correctness": "",
                    "reviewer_a_completeness": "",
                    "reviewer_a_error_grade": "",
                    "reviewer_a_recommendation_appropriate": "",
                    "reviewer_a_abstention_appropriate": "",
                    "reviewer_a_clinically_acceptable": "",
                    "reviewer_a_comment": "",
                    "reviewer_b_name": "",
                    "reviewer_b_correctness": "",
                    "reviewer_b_completeness": "",
                    "reviewer_b_error_grade": "",
                    "reviewer_b_recommendation_appropriate": "",
                    "reviewer_b_abstention_appropriate": "",
                    "reviewer_b_clinically_acceptable": "",
                    "reviewer_b_comment": "",
                    "adjudicator_name": "",
                    "adjudicated_correctness": "",
                    "adjudicated_completeness": "",
                    "adjudicated_error_grade": "",
                    "adjudicated_recommendation_appropriate": "",
                    "adjudicated_abstention_appropriate": "",
                    "adjudicated_clinically_acceptable": "",
                    "adjudication_comment": "",
                }
            )
        output.append(base)
    return output


def _simple_workbook(
    path: Path, sheet_name: str, rows: Sequence[dict[str, Any]], table: str
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    columns = list(rows[0]) if rows else ["status"]
    _append_dict_rows(sheet, list(rows) or [{"status": "pending"}], columns)
    _add_table(sheet, name=table)
    _set_widths(sheet, result_sheet=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_study_workbooks(
    *, root: Path, questions: Sequence[StudyQuestion]
) -> dict[str, Any]:
    rows, cells = build_planned_rows(root=root, questions=questions)
    excel_dir = root / "outputs/study_phase2/excel"
    excel_dir.mkdir(parents=True, exist_ok=True)
    manifest = json.loads(
        (root / "outputs/study_phase2/manifest/study_manifest.json").read_text(
            encoding="utf-8"
        )
    )
    refine = list(
        csv.DictReader(
            (root / "outputs/study_phase2/manifest/REFINE_compliance.csv").open(
                encoding="utf-8"
            )
        )
    )
    mi_clear = list(
        csv.DictReader(
            (root / "outputs/study_phase2/manifest/MI_CLEAR_LLM_compliance.csv").open(
                encoding="utf-8"
            )
        )
    )
    workbook = Workbook()
    readme = workbook.active
    readme.title = "00_README"
    readme_rows = [
        (
            "Zweck",
            "Reproduzierbarer Studienexport aus kanonischem JSONL; Fragen wurden per study_owner_pre_freeze_approval eingefroren, ohne unabhängige klinische Fragevalidierung.",
        ),
        ("WEB", "Orange: GPT mit verpflichtender Live-Websuche im jeweiligen Aufruf."),
        (
            "RAG",
            "Blau: GPT mit lokalem Closed-Corpus-Evidence-Package, ohne OpenAI-Tools.",
        ),
        (
            "Modelle",
            "GPT-5.5 grün; GPT-5.6 Sol violett. Unterschiedliche Reasoning-Efforts sind Teil der Deploymentkonfiguration.",
        ),
        ("Runs", "Run 1 kräftig = primärer Lauf; Run 2 hell = Reproduzierbarkeit."),
        (
            "Kosten",
            "Geschätzt = eingefrorene öffentliche Preise; kumulatives Fail-Closed-Limit 500 USD einschließlich Vorbereitung/Pilot/Retries; 400 USD ist historisch abgelöst.",
        ),
        (
            "Validierung",
            "Technische Provenienzvalidierung ist keine klinische Bewertung. Klinische Ratings bleiben leer bis zu zwei unabhängigen Reviews.",
        ),
        (
            "Einheiten",
            "Zeitfelder enden auf _ms; Kosten in USD mit sechs Dezimalstellen; Tokens sind ganzzahlig.",
        ),
        (
            "Filter",
            "Kopfzeile anklicken und Dropdownfilter nach Coverage, Modell, System, Run oder Status verwenden.",
        ),
        (
            "Sheets",
            "01 Manifest; 02 Fragen/Gold; 03 800 Ergebnisse; 04 Versuche; 05 Claims/Quellen; 06 Ressourcen; 07 Reproduzierbarkeit; 08 Ratings; 09 Adjudikation; 10 Statistik; 11 Compliance.",
        ),
    ]
    for row in readme_rows:
        readme.append(row)
    readme.column_dimensions["A"].width = 20
    readme.column_dimensions["B"].width = 110
    for row in readme.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    readme.freeze_panes = "A1"

    sheet = workbook.create_sheet("01_MANIFEST")
    flat_manifest = [
        {"Feld": key, "Wert": _json_cell(value)} for key, value in manifest.items()
    ]
    _append_dict_rows(sheet, flat_manifest, ["Feld", "Wert"])
    _add_table(sheet, name="StudyManifest")
    _set_widths(sheet, result_sheet=True)

    sheet = workbook.create_sheet("02_QUESTIONS_GOLD")
    question_rows = [row.model_dump(mode="json") for row in questions]
    q_columns = list(question_rows[0])
    serialized_q = [
        {key: _json_cell(value) for key, value in row.items()} for row in question_rows
    ]
    _append_dict_rows(sheet, serialized_q, q_columns)
    _add_table(sheet, name="QuestionsGold")
    _set_widths(sheet, result_sheet=True)

    sheet = workbook.create_sheet("03_ALL_RESULTS")
    _append_dict_rows(sheet, rows, RESULT_COLUMNS)
    _add_table(sheet, name="AllStudyResults")
    _set_widths(sheet, result_sheet=True)
    _style_results(sheet)

    attempts = read_jsonl(root / "outputs/study_phase2/results/api_attempts.jsonl")
    attempts_for_excel = [
        {key: _json_cell(value) for key, value in row.items()} for row in attempts
    ]
    sheet = workbook.create_sheet("04_API_ATTEMPTS")
    attempt_columns = list(attempts[0]) if attempts else ["attempt_id", "status"]
    _append_dict_rows(sheet, attempts_for_excel, attempt_columns)
    _add_table(sheet, name="ApiAttempts")
    _set_widths(sheet, result_sheet=True)

    sheet = workbook.create_sheet("05_CLAIMS_SOURCES")
    claim_rows: list[dict[str, Any]] = []
    for row in read_jsonl(root / "outputs/study_phase2/results/study_results.jsonl"):
        validated = row.get("validated_system_answer") or {}
        for claim in validated.get("claims") or ():
            claim_rows.append(
                {
                    "run_id": row["run_id"],
                    "question_id": row["question_id"],
                    "claim_id": claim.get("claim_id"),
                    "claim_text": claim.get("claim_text"),
                    "validator_status": claim.get("validator_status"),
                    "source_refs": _json_cell(claim.get("validated_source_refs")),
                }
            )
    claim_columns = list(claim_rows[0]) if claim_rows else ["run_id", "status"]
    _append_dict_rows(sheet, claim_rows, claim_columns)
    _add_table(sheet, name="ClaimsSources")
    _set_widths(sheet, result_sheet=True)

    analysis_specs = (
        (
            "06_RESOURCE_SUMMARY",
            root / "outputs/study_phase2/analysis/resource_summary.csv",
            {"Status": "pending", "Hinweis": "Nach Hauptstudie aus JSONL berechnen."},
        ),
        (
            "07_REPRODUCIBILITY",
            root / "outputs/study_phase2/analysis/reproducibility.csv",
            {"Status": "pending", "Hinweis": "Run-1/Run-2-Vergleich nach Hauptstudie."},
        ),
    )
    for name, analysis_path, placeholder in analysis_specs:
        if analysis_path.is_file():
            with analysis_path.open(encoding="utf-8", newline="") as handle:
                analysis_rows = list(csv.DictReader(handle))
        else:
            analysis_rows = [placeholder]
        columns = list(analysis_rows[0])
        sheet = workbook.create_sheet(name)
        _append_dict_rows(sheet, analysis_rows, columns)
        _add_table(sheet, name=name.replace("_", "")[:25])
        _set_widths(sheet, result_sheet=True)

    blind_rows = _blind_rows(rows, include_sources=False)
    sheet = workbook.create_sheet("08_RATINGS_BLINDED")
    _append_dict_rows(sheet, blind_rows, list(blind_rows[0]))
    _add_table(sheet, name="RatingsBlinded")
    _set_widths(sheet, result_sheet=True)

    sheet = workbook.create_sheet("09_ADJUDICATION")
    adjudication = [
        {
            "blinded_response_id": row["blinded_response_id"],
            "adjudicator": "",
            "final_error_grade": "",
            "final_clinically_acceptable": "",
            "comment": "",
        }
        for row in blind_rows
    ]
    _append_dict_rows(sheet, adjudication, list(adjudication[0]))
    _add_table(sheet, name="Adjudication")
    _set_widths(sheet, result_sheet=True)

    sheet = workbook.create_sheet("10_FINAL_STATISTICS")
    retrieval_summary_path = (
        root / "outputs/study_phase2/analysis/rag_retrieval_metrics.json"
    )
    if retrieval_summary_path.is_file():
        retrieval_summary = json.loads(
            retrieval_summary_path.read_text(encoding="utf-8")
        )
        final_rows = [
            {"Endpunkt": key, "Wert": _json_cell(value), "Status": "technical"}
            for key, value in retrieval_summary.items()
        ]
        final_rows.append(
            {
                "Endpunkt": "clinical_statistics",
                "Wert": "pending independent ratings, citation audit and adjudication",
                "Status": "pending",
            }
        )
    else:
        final_rows = [
            {
                "Endpunkt": "Status",
                "Wert": "pending_human_freeze_and_main_study",
                "Status": "pending",
            }
        ]
    _append_dict_rows(sheet, final_rows, list(final_rows[0]))
    _add_table(sheet, name="FinalStatistics")
    _set_widths(sheet, result_sheet=True)

    sheet = workbook.create_sheet("11_REFINE_MI_CLEAR")
    compliance = [{"Guideline": "REFINE", **row} for row in refine] + [
        {"Guideline": "MI-CLEAR-LLM", **row} for row in mi_clear
    ]
    columns = list(compliance[0])
    _append_dict_rows(sheet, compliance, columns)
    _add_table(sheet, name="ComplianceMatrix")
    _set_widths(sheet, result_sheet=True)

    master = excel_dir / "AISurgeon_RAG_vs_WEB_study_master.xlsx"
    workbook.save(master)

    arm_files = {
        ("gpt55_medium", "WEB"): "GPT55_MEDIUM_WEB.xlsx",
        ("gpt55_medium", "RAG"): "GPT55_MEDIUM_RAG.xlsx",
        ("gpt56_sol_high", "WEB"): "GPT56_SOL_HIGH_WEB.xlsx",
        ("gpt56_sol_high", "RAG"): "GPT56_SOL_HIGH_RAG.xlsx",
    }
    for (model, arm), filename in arm_files.items():
        subset = [
            row
            for row in rows
            if row["Modell"] == _display_model(model)
            and row["System_WEB_oder_RAG"] == arm
        ]
        if len(subset) != 200:
            raise AssertionError(f"{filename} expected 200 rows, got {len(subset)}")
        wb = Workbook()
        guide = wb.active
        guide.title = "00_README"
        guide.append(
            ["Inhalt", f"200 geplante Ergebnisse: {_display_model(model)}, {arm}."]
        )
        guide.append(
            ["Status", "Owner-Freeze erfolgt; technische/klinische Ergebnisfelder folgen aus dem Hauptlauf beziehungsweise späteren Ratings."]
        )
        guide.column_dimensions["A"].width = 18
        guide.column_dimensions["B"].width = 90
        result_sheet = wb.create_sheet("03_ALL_RESULTS")
        _append_dict_rows(result_sheet, subset, RESULT_COLUMNS)
        _add_table(result_sheet, name=f"Results{model}{arm}")
        _set_widths(result_sheet, result_sheet=True)
        _style_results(result_sheet)
        wb.save(excel_dir / filename)

    rating_dir = root / "outputs/study_phase2/ratings"
    _simple_workbook(
        rating_dir / "clinical_ratings_blinded.xlsx",
        "RATINGS_BLINDED",
        blind_rows,
        "ClinicalRatings",
    )
    citation_rows = _blind_rows(rows, include_sources=True)
    for row in citation_rows:
        row.update(
            {
                "citation_reviewer_name": "",
                "source_exists": "",
                "source_quality": "",
                "claim_source_support": "",
                "citation_completeness": "",
                "locator_correct": "",
                "invented_source": "",
                "direct_vs_relation_correct": "",
                "reviewer_comment": "",
            }
        )
    _simple_workbook(
        rating_dir / "citation_audit.xlsx",
        "CITATION_AUDIT",
        citation_rows,
        "CitationAudit",
    )
    return {
        "master": str(master),
        "arm_files": [str(excel_dir / value) for value in arm_files.values()],
        "review": str(
            root / "outputs/study_phase2/questions/question_freeze_review.xlsx"
        ),
        "planned_rows": len(rows),
        "cells": len(cells),
    }


def validate_study_workbooks(*, root: Path) -> dict[str, Any]:
    excel_dir = root / "outputs/study_phase2/excel"
    expected_master_sheets = [
        "00_README",
        "01_MANIFEST",
        "02_QUESTIONS_GOLD",
        "03_ALL_RESULTS",
        "04_API_ATTEMPTS",
        "05_CLAIMS_SOURCES",
        "06_RESOURCE_SUMMARY",
        "07_REPRODUCIBILITY",
        "08_RATINGS_BLINDED",
        "09_ADJUDICATION",
        "10_FINAL_STATISTICS",
        "11_REFINE_MI_CLEAR",
    ]
    paths = [
        excel_dir / "AISurgeon_RAG_vs_WEB_study_master.xlsx",
        excel_dir / "GPT55_MEDIUM_WEB.xlsx",
        excel_dir / "GPT55_MEDIUM_RAG.xlsx",
        excel_dir / "GPT56_SOL_HIGH_WEB.xlsx",
        excel_dir / "GPT56_SOL_HIGH_RAG.xlsx",
    ]
    checks: list[dict[str, Any]] = []
    for index, path in enumerate(paths):
        workbook = load_workbook(path, data_only=False, read_only=False)
        if index == 0 and workbook.sheetnames != expected_master_sheets:
            raise AssertionError(
                f"unexpected master sheet names: {workbook.sheetnames}"
            )
        sheet = workbook["03_ALL_RESULTS"]
        expected_rows = 800 if index == 0 else 200
        if sheet.max_row - 1 != expected_rows:
            raise AssertionError(f"{path.name}: expected {expected_rows} rows")
        headers = {cell.value: cell.column for cell in sheet[1]}
        run_ids = [
            sheet.cell(row, headers["Run_ID"]).value
            for row in range(2, sheet.max_row + 1)
        ]
        if len(run_ids) != len(set(run_ids)):
            raise AssertionError(f"{path.name}: duplicate run IDs")
        for required in RESULT_COLUMNS[:24]:
            if required not in headers:
                raise AssertionError(f"{path.name}: missing {required}")
        if sheet.freeze_panes != "A2" or not sheet.auto_filter.ref:
            raise AssertionError(f"{path.name}: filters/freeze panes missing")
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#REF!"):
                    raise AssertionError(f"{path.name}: reference error")
        checks.append(
            {
                "file": str(path),
                "sheets": workbook.sheetnames,
                "result_rows": expected_rows,
                "unique_run_ids": len(run_ids),
                "freeze_panes": str(sheet.freeze_panes),
                "autofilter": sheet.auto_filter.ref,
                "status": "passed",
            }
        )
    return {"status": "passed", "files": checks}


__all__ = [
    "RESULT_COLUMNS",
    "build_question_review_workbook",
    "build_study_workbooks",
    "export_planned_results",
    "validate_study_workbooks",
]
