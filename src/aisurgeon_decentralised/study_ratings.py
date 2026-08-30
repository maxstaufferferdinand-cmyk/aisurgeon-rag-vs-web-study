"""Idempotent import of independent clinical ratings and citation audit."""

from __future__ import annotations

import json
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .study_analysis import build_artifact_hash_manifest, write_flat_csv
from .study_exports import _add_table, _append_dict_rows, _set_widths
from .study_phase2 import read_jsonl, sha256_file, utc_now
from .study_statistics import cluster_bootstrap_mean_ci

ERROR_GRADES = {
    "none": 0,
    "no_error": 0,
    "kein_fehler": 0,
    "minor": 1,
    "minor_error": 1,
    "major": 2,
    "major_error": 2,
    "critical": 3,
    "critical_error": 3,
    "potentially_harmful": 3,
    "potenziell_schaedlich": 3,
}
TRUE_VALUES = {"true", "yes", "ja", "1", "x", "acceptable", "angemessen"}
FALSE_VALUES = {"false", "no", "nein", "0", "not_acceptable", "nicht_angemessen"}
NOT_APPLICABLE_VALUES = {"n/a", "na", "not_applicable", "nicht_anwendbar"}


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def _sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "") for cell in sheet[1]]
    if not headers or any(not header for header in headers):
        raise RuntimeError(f"{path.name}/{sheet_name}: invalid header row")
    rows = [
        dict(zip(headers, values, strict=True))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]
    return [row for row in rows if any(value not in (None, "") for value in row.values())]


def _nonempty(row: dict[str, Any], fields: tuple[str, ...], row_id: str) -> None:
    missing = [field for field in fields if not str(row.get(field) or "").strip()]
    if missing:
        raise RuntimeError(f"{row_id}: missing rating fields {missing}")


def _grade(value: Any) -> int:
    normalized = str(value or "").strip().casefold().replace(" ", "_")
    if normalized not in ERROR_GRADES:
        raise RuntimeError(f"invalid error grade: {value!r}")
    return ERROR_GRADES[normalized]


def _boolean(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().casefold().replace(" ", "_")
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    raise RuntimeError(f"invalid yes/no rating: {value!r}")


def _optional_boolean(value: Any) -> bool | None:
    normalized = str(value or "").strip().casefold().replace(" ", "_")
    if normalized in NOT_APPLICABLE_VALUES:
        return None
    return _boolean(value)


def weighted_kappa(left: list[int], right: list[int], *, categories: int = 4) -> float:
    if len(left) != len(right) or not left:
        raise ValueError("weighted kappa requires equally sized nonempty ratings")
    matrix = [[0.0 for _ in range(categories)] for _ in range(categories)]
    for a, b in zip(left, right, strict=True):
        matrix[a][b] += 1.0
    total = float(len(left))
    left_marginal = [sum(row) for row in matrix]
    right_marginal = [sum(matrix[i][j] for i in range(categories)) for j in range(categories)]
    observed = 0.0
    expected = 0.0
    denominator = float((categories - 1) ** 2)
    for i in range(categories):
        for j in range(categories):
            weight = ((i - j) ** 2) / denominator
            observed += weight * matrix[i][j] / total
            expected += weight * (left_marginal[i] * right_marginal[j]) / (total * total)
    return 1.0 - observed / expected if expected else 1.0


def _replace_sheet(workbook: Any, name: str, rows: list[dict[str, Any]], table_name: str) -> None:
    if name in workbook.sheetnames:
        index = workbook.sheetnames.index(name)
        workbook.remove(workbook[name])
        sheet = workbook.create_sheet(name, index)
    else:
        sheet = workbook.create_sheet(name)
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    if not columns:
        columns = ["status"]
    _append_dict_rows(sheet, rows or [{"status": "pending"}], columns)
    _add_table(sheet, name=table_name)
    _set_widths(sheet, result_sheet=True)


def import_human_ratings(*, root: Path) -> dict[str, Any]:
    root = root.resolve()
    base = root / "outputs/study_phase2"
    clinical_path = base / "ratings/clinical_ratings_blinded.xlsx"
    citation_path = base / "ratings/citation_audit.xlsx"
    source_hashes = {
        "clinical_ratings_sha256": sha256_file(clinical_path),
        "citation_audit_sha256": sha256_file(citation_path),
    }
    final_report_path = base / "reports/final_clinical_report.json"
    if final_report_path.is_file():
        existing = json.loads(final_report_path.read_text(encoding="utf-8"))
        if existing.get("rating_source_hashes") == source_hashes:
            return existing

    clinical = _sheet_rows(clinical_path, "RATINGS_BLINDED")
    citation = _sheet_rows(citation_path, "CITATION_AUDIT")
    if len(clinical) != 400 or len(citation) != 400:
        raise RuntimeError(
            f"ratings incomplete: clinical={len(clinical)}, citation={len(citation)}; expected 400 each"
        )
    clinical_ids = [str(row["blinded_response_id"]) for row in clinical]
    citation_ids = [str(row["blinded_response_id"]) for row in citation]
    if len(set(clinical_ids)) != 400 or set(clinical_ids) != set(citation_ids):
        raise RuntimeError("rating files have duplicate or mismatched blinded IDs")

    required_by_reviewer = (
        "correctness",
        "completeness",
        "error_grade",
        "recommendation_appropriate",
        "abstention_appropriate",
        "clinically_acceptable",
    )
    reviewer_a_names: set[str] = set()
    reviewer_b_names: set[str] = set()
    enriched: list[dict[str, Any]] = []
    for row in clinical:
        row_id = str(row["blinded_response_id"])
        required = (
            "reviewer_a_name",
            "reviewer_b_name",
            "adjudicator_name",
            "adjudicated_correctness",
            "adjudicated_completeness",
            "adjudicated_error_grade",
            "adjudicated_recommendation_appropriate",
            "adjudicated_abstention_appropriate",
            "adjudicated_clinically_acceptable",
            *(f"reviewer_a_{field}" for field in required_by_reviewer),
            *(f"reviewer_b_{field}" for field in required_by_reviewer),
        )
        _nonempty(row, required, row_id)
        reviewer_a_names.add(str(row["reviewer_a_name"]).strip())
        reviewer_b_names.add(str(row["reviewer_b_name"]).strip())
        enriched.append(
            {
                **row,
                "reviewer_a_error_grade_ordinal": _grade(
                    row["reviewer_a_error_grade"]
                ),
                "reviewer_b_error_grade_ordinal": _grade(
                    row["reviewer_b_error_grade"]
                ),
                "adjudicated_error_grade_ordinal": _grade(
                    row["adjudicated_error_grade"]
                ),
                "reviewer_a_clinically_acceptable_bool": _boolean(
                    row["reviewer_a_clinically_acceptable"]
                ),
                "reviewer_b_clinically_acceptable_bool": _boolean(
                    row["reviewer_b_clinically_acceptable"]
                ),
                "adjudicated_clinically_acceptable_bool": _boolean(
                    row["adjudicated_clinically_acceptable"]
                ),
                "adjudicated_recommendation_appropriate_bool": _optional_boolean(
                    row["adjudicated_recommendation_appropriate"]
                ),
                "adjudicated_abstention_appropriate_bool": _optional_boolean(
                    row["adjudicated_abstention_appropriate"]
                ),
            }
        )
    if reviewer_a_names.intersection(reviewer_b_names):
        raise RuntimeError("Reviewer A and B identities must be independent")

    citation_required = (
        "citation_reviewer_name",
        "source_exists",
        "source_quality",
        "claim_source_support",
        "citation_completeness",
        "locator_correct",
        "invented_source",
        "direct_vs_relation_correct",
    )
    for row in citation:
        _nonempty(row, citation_required, str(row["blinded_response_id"]))

    citation_by_id = {str(row["blinded_response_id"]): row for row in citation}
    results = {
        str(row["run_id"]): row
        for row in read_jsonl(base / "results/study_results.jsonl")
    }
    imported: list[dict[str, Any]] = []
    for row in enriched:
        citation_row = citation_by_id[str(row["blinded_response_id"])]
        run_id = str(citation_row["run_id"])
        result = results.get(run_id)
        if result is None:
            raise RuntimeError(f"citation audit references unknown run_id: {run_id}")
        imported.append(
            {
                **row,
                "run_id": run_id,
                "question_id": result["question_id"],
                "coverage_stratum": result["coverage_stratum"],
                "model_config_id": result["model_config_id"],
                "system_arm": result["system_arm"],
            }
        )
    write_flat_csv(base / "ratings/clinical_ratings_imported.csv", imported)
    write_flat_csv(base / "ratings/citation_audit_imported.csv", citation)
    from .study_phase2 import write_jsonl_atomic

    write_jsonl_atomic(base / "ratings/clinical_ratings_imported.jsonl", imported)
    write_jsonl_atomic(base / "ratings/citation_audit_imported.jsonl", citation)

    group_rows: list[dict[str, Any]] = []
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in imported:
        grouped[
            (
                str(row["model_config_id"]),
                str(row["system_arm"]),
                str(row["coverage_stratum"]),
            )
        ].append(row)
    for (model, arm, stratum), rows in sorted(grouped.items()):
        abstention_a = [
            value
            for row in rows
            if (
                value := _optional_boolean(
                    row["reviewer_a_abstention_appropriate"]
                )
            )
            is not None
        ]
        abstention_b = [
            value
            for row in rows
            if (
                value := _optional_boolean(
                    row["reviewer_b_abstention_appropriate"]
                )
            )
            is not None
        ]
        group_rows.append(
            {
                "analysis_type": "model_arm_coverage_summary",
                "model_config_id": model,
                "system_arm": arm,
                "coverage_stratum": stratum,
                "n": len(rows),
                "clinically_acceptable_rate": statistics.mean(
                    float(row["adjudicated_clinically_acceptable_bool"])
                    for row in rows
                ),
                "major_or_critical_error_rate": statistics.mean(
                    float(row["adjudicated_error_grade_ordinal"] >= 2)
                    for row in rows
                ),
                "critical_error_rate": statistics.mean(
                    float(row["adjudicated_error_grade_ordinal"] >= 3)
                    for row in rows
                ),
                "appropriate_abstention_rate_reviewer_a": statistics.mean(
                    float(value) for value in abstention_a
                )
                if abstention_a
                else None,
                "appropriate_abstention_rate_reviewer_b": statistics.mean(
                    float(value) for value in abstention_b
                )
                if abstention_b
                else None,
                "adjudicated_appropriate_abstention_rate": statistics.mean(
                    float(value)
                    for row in rows
                    if (
                        value := row[
                            "adjudicated_abstention_appropriate_bool"
                        ]
                    )
                    is not None
                )
                if any(
                    row["adjudicated_abstention_appropriate_bool"] is not None
                    for row in rows
                )
                else None,
                "cost_per_clinically_acceptable_answer_usd": (
                    sum(
                        float(
                            (results[str(row["run_id"])].get("cost") or {}).get(
                                "total_estimated_cost_usd"
                            )
                            or 0
                        )
                        for row in rows
                    )
                    / sum(
                        bool(row["adjudicated_clinically_acceptable_bool"])
                        for row in rows
                    )
                    if any(
                        row["adjudicated_clinically_acceptable_bool"] for row in rows
                    )
                    else None
                ),
            }
        )

    paired_rows: list[dict[str, Any]] = []
    by_cell = {
        (
            str(row["question_id"]),
            str(row["model_config_id"]),
            str(row["system_arm"]),
        ): row
        for row in imported
    }
    models = sorted({str(row["model_config_id"]) for row in imported})
    for model in models:
        for stratum in (
            "all_prespecified_80_20",
            "covered_by_local_corpus",
            "not_covered_by_local_corpus",
        ):
            question_ids = sorted(
                {
                    str(row["question_id"])
                    for row in imported
                    if row["model_config_id"] == model
                    and (
                        stratum == "all_prespecified_80_20"
                        or row["coverage_stratum"] == stratum
                    )
                }
            )
            endpoints = {
                "clinically_acceptable": lambda row: float(
                    row["adjudicated_clinically_acceptable_bool"]
                ),
                "major_or_critical_error": lambda row: float(
                    row["adjudicated_error_grade_ordinal"] >= 2
                ),
                "critical_error": lambda row: float(
                    row["adjudicated_error_grade_ordinal"] >= 3
                ),
                "appropriate_abstention": lambda row: (
                    None
                    if row["adjudicated_abstention_appropriate_bool"] is None
                    else float(row["adjudicated_abstention_appropriate_bool"])
                ),
            }
            for endpoint, extractor in endpoints.items():
                differences: dict[str, float] = {}
                for question_id in question_ids:
                    rag = by_cell.get((question_id, model, "RAG"))
                    web = by_cell.get((question_id, model, "WEB"))
                    if rag is None or web is None:
                        continue
                    rag_value = extractor(rag)
                    web_value = extractor(web)
                    if rag_value is None or web_value is None:
                        continue
                    differences[question_id] = rag_value - web_value
                ci_low, ci_high = cluster_bootstrap_mean_ci(
                    differences,
                    seed=20260829 + len(paired_rows),
                )
                paired_rows.append(
                    {
                        "analysis_type": "paired_rag_minus_web",
                        "model_config_id": model,
                        "coverage_stratum": stratum,
                        "endpoint": endpoint,
                        "questions_paired": len(differences),
                        "mean_difference_rag_minus_web": statistics.mean(
                            differences.values()
                        )
                        if differences
                        else None,
                        "ci95_low": ci_low,
                        "ci95_high": ci_high,
                        "bootstrap_resamples": 10_000,
                        "cluster_unit": "question_id",
                    }
                )
    all_clinical_statistics = [*group_rows, *paired_rows]
    write_flat_csv(
        base / "analysis/clinical_statistics.csv", all_clinical_statistics
    )

    kappa = weighted_kappa(
        [int(row["reviewer_a_error_grade_ordinal"]) for row in imported],
        [int(row["reviewer_b_error_grade_ordinal"]) for row in imported],
    )
    acceptable_agreement = statistics.mean(
        float(
            row["reviewer_a_clinically_acceptable_bool"]
            == row["reviewer_b_clinically_acceptable_bool"]
        )
        for row in imported
    )
    final = {
        "schema_version": "phase2-final-clinical-report-1.0.0",
        "created_at_utc": utc_now(),
        "status": "CLINICAL_STUDY_COMPLETE",
        "rating_source_hashes": source_hashes,
        "clinical_ratings": len(imported),
        "citation_audits": len(citation),
        "reviewer_a_identities": sorted(reviewer_a_names),
        "reviewer_b_identities": sorted(reviewer_b_names),
        "quadratic_weighted_kappa_error_grade": kappa,
        "raw_agreement_clinically_acceptable": acceptable_agreement,
        "adjudication_completed": True,
        "group_statistics": group_rows,
        "paired_system_statistics": paired_rows,
        "citation_technical_counts": {
            field: dict(sorted(Counter(str(row[field]) for row in citation).items()))
            for field in citation_required[1:]
        },
        "interpretation_scope": (
            "Synthetic prespecified 80/20 benchmark; not an estimate of real-world "
            "coverage prevalence or standalone clinical-device validation."
        ),
    }
    _write_json(final_report_path, final)
    (base / "reports/final_clinical_report.md").write_text(
        "\n".join(
            [
                "# Finaler klinischer Studienbericht",
                "",
                f"Status: `{final['status']}`",
                "",
                (
                    "400 Run-1-Antworten wurden von mindestens zwei unabhängigen "
                    "Reviewern bewertet und adjudiziert; 400 Citation Audits wurden "
                    "importiert. Quadratisch gewichtetes Kappa der Fehlergrade: "
                    f"{kappa:.4f}."
                ),
                "",
                (
                    "Die Ergebnisse beziehen sich auf das synthetische, "
                    "prä-spezifizierte 80/20-Benchmarkgemisch und belegen keine "
                    "allgemeine klinische Sicherheit oder reale Coverage-Prävalenz."
                ),
                "",
            ]
        ),
        encoding="utf-8",
        newline="\n",
    )

    master_path = base / "excel/AISurgeon_RAG_vs_WEB_study_master.xlsx"
    workbook = load_workbook(master_path)
    _replace_sheet(workbook, "08_RATINGS_BLINDED", clinical, "RatingsImported")
    adjudication_rows = [
        {
            "blinded_response_id": row["blinded_response_id"],
            "adjudicator_name": row["adjudicator_name"],
            "adjudicated_error_grade": row["adjudicated_error_grade"],
            "adjudicated_correctness": row["adjudicated_correctness"],
            "adjudicated_completeness": row["adjudicated_completeness"],
            "adjudicated_recommendation_appropriate": row[
                "adjudicated_recommendation_appropriate"
            ],
            "adjudicated_abstention_appropriate": row[
                "adjudicated_abstention_appropriate"
            ],
            "adjudicated_clinically_acceptable": row[
                "adjudicated_clinically_acceptable"
            ],
        }
        for row in clinical
    ]
    _replace_sheet(workbook, "09_ADJUDICATION", adjudication_rows, "Adjudicated")
    _replace_sheet(
        workbook,
        "10_FINAL_STATISTICS",
        all_clinical_statistics,
        "ClinicalStatistics",
    )
    workbook.save(master_path)

    manifest_path = base / "manifest/study_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update(
        {
            "status": "CLINICAL_STUDY_COMPLETE",
            "clinical_rating_status": "complete",
            "rating_source_hashes": source_hashes,
            "clinical_report_sha256": sha256_file(final_report_path),
        }
    )
    _write_json(manifest_path, manifest)
    hashes = build_artifact_hash_manifest(
        root=root,
        paths=[
            clinical_path,
            citation_path,
            final_report_path,
            base / "reports/final_clinical_report.md",
            master_path,
            manifest_path,
        ],
    )
    _write_json(base / "manifest/artifact_hashes_clinical_complete.json", hashes)
    return final


__all__ = ["import_human_ratings", "weighted_kappa"]
